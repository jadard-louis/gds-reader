#!/usr/bin/env python3
"""
GDS Reader  —  Local Flask API Server
啟動後在瀏覽器開啟 http://localhost:5000
"""

import os, sys, json, tempfile, traceback
sys.setrecursionlimit(10000)
from pathlib import Path
from collections import defaultdict
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS

# 確保同目錄模組可被 import（無論從哪個工作目錄啟動）
_HERE = Path(__file__).parent.resolve()
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

# ── 分析邏輯（不依賴 tkinter）────────────────────────────────────────────────
PURPOSES = ["bump", "amark", "amark_text", "bump_text_name", "bump_text_number", "chip_size", "origin_mark", "unknown"]
PURPOSE_LABELS = {
    "bump":             "Bump / PAD",
    "amark":            "Alignment Mark",
    "amark_text":       "Alignment Mark Text",
    "bump_text_name":   "Bump Text (名稱)",
    "bump_text_number": "Bump Text (編號)",
    "chip_size":        "Chip Size",
    "origin_mark":      "Origin Mark (0,0)",
    "unknown":          "Unknown",
}

def _feat(pts):
    import numpy as np
    pts = np.array(pts)
    w = float(pts[:,0].max() - pts[:,0].min())
    h = float(pts[:,1].max() - pts[:,1].min())
    trapz = getattr(np, "trapezoid", None) or getattr(np, "trapz")
    a = abs(float(trapz(pts[:,1], pts[:,0])))
    return {"w": w, "h": h, "area": a, "n": len(pts)}

# Layer 編號 → 強制用途（只適用 bump 等明確層）
LAYER_OVERRIDES = {
    204: "bump",
    224: "bump",   # 大量 polygon = Bump/PAD
    225: "bump",   # Bump/PAD
}

# (layer, datatype) 精確對應（優先於 LAYER_OVERRIDES）
LAYER_DTYPE_OVERRIDES = {
    (190, 9): "chip_size",   # 190:9 預設晶片邊界層
}

# Layer 編號 → 文字強制用途（None = 交由 _guess 內容自動偵測）
LAYER_TEXT_OVERRIDES = {
    225: None,   # 225 文字含 NO → 編號，交由內容偵測
}

def _guess(lyr, dt, polys, texts, scale, ctx_polys=None):
    import numpy as np
    import sys as _sys_dbg_guess

    # 0. (layer, datatype) 精確強制規則（最高優先）
    if polys and (lyr, dt) in LAYER_DTYPE_OVERRIDES:
        return LAYER_DTYPE_OVERRIDES[(lyr, dt)]

    # 0.5. Layer 224, 225 DType 0 強制識別規則（不受推測影響）
    #      - 帶有 #POLY → 一律為 BUMP/PAD
    #      - 帶有 #TEXT → 一律為 Bump 序號標注
    if (lyr, dt) in [(224, 0), (225, 0)]:
        if polys:
            print(f"[FORCE_RULE] Layer {lyr} DType {dt}: {len(polys)} polys → BUMP", file=_sys_dbg_guess.stderr)
            return "bump"  # Layer 224, 225 DType 0 有 polygon 一律為 BUMP/PAD
        elif texts:
            print(f"[FORCE_RULE] Layer {lyr} DType {dt}: {len(texts)} texts → bump_text_number", file=_sys_dbg_guess.stderr)
            return "bump_text_number"  # Layer 224, 225 DType 0 有 text 一律為 Bump 序號

    # 1. layer 強制規則（只在有 polygon 時套用）
    if polys and lyr in LAYER_OVERRIDES:
        return LAYER_OVERRIDES[lyr]

    # 2. 有多邊形 → 幾何判斷
    if polys:
        feats  = [_feat(p.points) for p in polys]
        areas  = [f["area"] * scale * scale for f in feats]
        ws     = [f["w"] * scale for f in feats]
        hs     = [f["h"] * scale for f in feats]
        npts   = [f["n"] for f in feats]
        avg_a  = float(np.mean(areas))
        avg_ar = float(np.mean([w/h if h > 0 else 0 for w, h in zip(ws, hs)]))
        max_n  = int(max(npts))
        min_n  = int(min(npts))
        max_w  = float(max(ws))
        max_h  = float(max(hs))
        cnt    = len(polys)

        # Chip size: 極少多邊形 + 超大面積
        if cnt <= 4 and avg_a > 1e5:
            return "chip_size"

        # 原點標記：極少矩形 + 極小面積（10×10 μm，< 150 μm²）
        if cnt <= 4 and max_n == 4 and avg_a < 150 and max_w < 15 and max_h < 15:
            return "origin_mark"

        # Amark 偵測（205~211 等 Alignment Mark 相關層）
        if max_w <= 200 and max_h <= 200:
            # Amark 外框矩形：≤4 個矩形，尺寸 80~150 μm（面積 5000~25000 μm²）
            if cnt <= 4 and max_n == 4 and 5000 <= avg_a <= 25000 and 0.5 < avg_ar < 2.0:
                return "amark"
            # 十字形本體（12pts）或含矩形（4pts），最多 8 個
            if cnt <= 8 and max_n >= 8:
                return "amark"
            # 三角形樣式（3pts each）+ 外框矩形
            if cnt <= 20 and min_n == 3:
                return "amark"

        # Bump / PAD：數量多、面積中等、長寬比近正方形
        if cnt > 4 and avg_a < 5000 and 0.5 < avg_ar < 2.0:
            return "bump"

    # 3. text 行：先偵測 ctx_polys 屬性，再判斷文字內容
    if texts:
        if ctx_polys:
            ctx_feats = [_feat(p.points) for p in ctx_polys]
            ctx_areas = [f["area"] * scale * scale for f in ctx_feats]
            ctx_npts  = [f["n"] for f in ctx_feats]
            ctx_avg_a = float(np.mean(ctx_areas))
            ctx_max_n = int(max(ctx_npts))
            # ctx polygons 為 amark 外框矩形 → 文字歸 amark_text
            if len(ctx_polys) <= 4 and ctx_max_n == 4 and 5000 <= ctx_avg_a <= 25000:
                return "amark_text"
            # ctx polygons 面積過大（amark 十字 / chip_size 附屬）→ 忽略
            if ctx_avg_a > 10000:
                return "unknown"
        # 文字內容偵測
        if any("NO" in t.upper() for t in texts):
            return "bump_text_number"
        has_alpha = any(any(c.isalpha() for c in t) for t in texts)
        has_digit = any(t.isdigit() for t in texts)
        return "bump_text_name" if has_alpha else ("bump_text_number" if has_digit else "bump_text_name")

    return "unknown"

def analyze_gds(path):
    import gdstk, numpy as np
    lib   = gdstk.read_gds(path)
    scale = lib.unit / 1e-6

    # 使用 top-level 絕對座標（flatten 展平所有 reference 位移/旋轉）
    tops = lib.top_level()
    if not tops:
        raise ValueError("No top-level cell found")
    top_flat = tops[0].copy("_analyze_flat", deep_copy=True)
    top_flat.flatten()

    layer_polys = defaultdict(list)
    layer_texts = defaultdict(list)
    for poly in top_flat.polygons:
        layer_polys[(poly.layer, poly.datatype)].append(poly)
    for label in top_flat.labels:
        layer_texts[(label.layer, label.texttype)].append(label.text)

    rows = []
    for (lyr, dt) in sorted(set(layer_polys) | set(layer_texts)):
        polys = layer_polys[(lyr, dt)]
        texts = layer_texts[(lyr, dt)]

        # 同一 layer 同時有 polygon 與 text → 一律拆成兩行分別處理
        if polys and texts:
            poly_groups = [(polys, [], None), ([], texts, polys)]
        else:
            poly_groups = [(polys, texts, None)]

        for grp_polys, grp_texts, ctx_polys in poly_groups:
            purpose = _guess(lyr, dt, grp_polys, grp_texts, scale, ctx_polys)
            # text 行：若在 LAYER_TEXT_OVERRIDES 有強制覆寫則套用
            if grp_texts and not grp_polys and lyr in LAYER_TEXT_OVERRIDES:
                override = LAYER_TEXT_OVERRIDES[lyr]
                if override is not None:
                    purpose = override

            feats = [_feat(p.points) for p in grp_polys] if grp_polys else []
            ws    = [f["w"] * scale for f in feats]
            hs    = [f["h"] * scale for f in feats]
            areas = [f["area"] * scale * scale for f in feats]
            rows.append({
                "layer":    lyr,
                "datatype": dt,
                "n_poly":   len(grp_polys),
                "n_text":   len(grp_texts),
                "avg_w":    round(float(np.mean(ws)),   3) if ws    else 0,
                "avg_h":    round(float(np.mean(hs)),   3) if hs    else 0,
                "avg_area": round(float(np.mean(areas)),2) if areas else 0,
                "purpose":  purpose,
            })

    tops = [c.name for c in lib.top_level()]
    info = {
        "filename": Path(path).name,
        "unit":     lib.unit,
        "cells":    len(lib.cells),
        "tops":     tops,
        "layers":   len(rows),
    }
    return rows, info

app = Flask(__name__, static_folder=str(Path(__file__).parent))
CORS(app)   # 允許瀏覽器跨來源呼叫

# 暫存上傳檔案的路徑與原始檔名（本次 session 用）
_uploaded_gds_path: str = ""
_uploaded_gds_name: str = ""

# ── 靜態頁面 ──────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(app.static_folder, "gds_gui_preview.html")


# ── 輸出檔案瀏覽（供瀏覽器直接開啟 PNG / Excel）────────────────────────────────
@app.route("/output/<path:filename>")
def serve_output(filename):
    out_dir = str(_HERE / "output")
    return send_from_directory(out_dir, filename)


# ── 診斷：列出輸出目錄中的所有檔案 ─────────────────────────────────────────────
@app.route("/api/list_output_files", methods=["GET"])
def list_output_files():
    """列出 output 目錄中的所有檔案（用於診斷）"""
    out_dir = _HERE / "output"
    if not out_dir.exists():
        return jsonify({"error": "Output directory does not exist", "files": []})

    try:
        files = []
        for f in sorted(out_dir.glob("*")):
            if f.is_file():
                files.append({
                    "name": f.name,
                    "size": f.stat().st_size,
                    "path": str(f),
                    "url": f"/output/{f.name}"
                })
        return jsonify({"output_dir": str(out_dir), "files": files, "count": len(files)})
    except Exception as e:
        return jsonify({"error": str(e), "files": []}), 500


# ── 診斷：保存日誌檔案 ───────────────────────────────────────────────────────────
# ── 診斷：打開輸出資料夾 ───────────────────────────────────────────────────────
@app.route("/api/open_folder", methods=["POST"])
def open_folder():
    """打開輸出資料夾讓用戶手動選擇檔案"""
    body = request.json or {}
    out_dir = body.get("out_dir", "").strip() or str(_HERE / "output")

    try:
        out_path = Path(out_dir)
        if not out_path.exists():
            return jsonify({"error": "資料夾不存在"}), 400

        if os.environ.get("RENDER"):
            return jsonify({"message": "雲端模式：請直接下載檔案，無法開啟本機資料夾"})

        import subprocess
        import platform

        # 根據作業系統打開資料夾
        if platform.system() == "Windows":
            subprocess.Popen(f'explorer "{out_path}"')
        elif platform.system() == "Darwin":  # macOS
            subprocess.Popen(["open", str(out_path)])
        else:  # Linux
            subprocess.Popen(["xdg-open", str(out_path)])

        return jsonify({"message": f"已打開資料夾: {out_path}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── 上傳 GDS 檔案（拖曳或 Browse 後上傳二進位）────────────────────────────────
@app.route("/api/upload_gds", methods=["POST"])
def upload_gds():
    global _uploaded_gds_path, _uploaded_gds_name
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "沒有收到檔案"}), 400

    suffix = Path(f.filename).suffix or ".gds"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix,
                                     prefix="gds_upload_")
    f.save(tmp.name)
    _uploaded_gds_path = tmp.name
    _uploaded_gds_name = Path(f.filename).stem   # 原始檔名（不含副檔名）
    return jsonify({"path": f.filename, "tmp_path": tmp.name})


def _project_name(stem: str) -> str:
    """取檔名第一個 '_' 前的字串作為 project name（例如 JD9522TS_CB6_... → JD9522TS）"""
    return stem.split("_")[0] if "_" in stem else stem

# ── 掃描 GDS ──────────────────────────────────────────────────────────────────
@app.route("/api/scan", methods=["POST"])
def api_scan():
    body = request.json or {}
    path = body.get("file_path", "").strip()

    # 如果前端給的是剛上傳的暫存路徑就直接用，否則嘗試本機路徑
    if not path and _uploaded_gds_path:
        path = _uploaded_gds_path
    if not path:
        return jsonify({"error": "請提供 GDS 檔案路徑"}), 400
    if not Path(path).exists():
        return jsonify({"error": f"找不到檔案：{path}"}), 404

    try:
        rows, info = analyze_gds(path)
        if isinstance(info, str):
            return jsonify({"error": info}), 500

        # ── 檢查是否有已儲存的 mapping JSON（僅用於提示，不覆蓋推測結果） ──────────────
        stem    = _uploaded_gds_name if _uploaded_gds_name else Path(path).stem
        pname   = _project_name(stem)
        out_dir = str(_HERE / "output")
        mapping_json = Path(out_dir) / f"{pname}.layer_mapping.json"
        mapping_loaded = False
        mapping_path = None
        if mapping_json.exists():
            mapping_loaded = True
            mapping_path = str(mapping_json)

        # 轉成前端期望的欄位名稱；不自動應用已儲存的 mapping，由用戶明確點擊「加載 Mapping JSON」時才覆寫
        table_rows = []
        for r in rows:
            purpose = r["purpose"]  # 直接使用掃描推測結果，不覆寫
            table_rows.append({
                "layer":   r["layer"],
                "dt":      r["datatype"],
                "nPoly":   r["n_poly"],
                "nText":   r["n_text"],
                "w":       r["avg_w"],
                "h":       r["avg_h"],
                "area":    r["avg_area"],
                "purpose": purpose,
            })

        # 計算 Bump Type 統計（掃描時快速計算）
        import importlib, sys as _sys
        try:
            if "test_export_xlsx" in _sys.modules:
                _mod = importlib.reload(_sys.modules["test_export_xlsx"])
            else:
                _mod = importlib.import_module("test_export_xlsx")

            # 注意：掃描時不自動套用 mapping JSON，使用預設層配置
            # Bump Type 計算基於掃描後立即推測的結果，由用戶明確加載 mapping 時才覆寫

            _top, _scale = _mod.load_flat(path)
            _bumps = _mod.get_bumps(_top, _scale)

            # 空間分類
            _chip_bbox = _mod.get_chip_bbox(_top, _scale)
            _chip_cy = (_chip_bbox[1] + _chip_bbox[3]) / 2 if _chip_bbox else 0
            _y_grps = {}
            for _b in _bumps:
                _ky = round(_b["cy"] * 2) / 2
                _y_grps.setdefault(_ky, []).append(id(_b))

            _input_ids = {i for ky, ids in _y_grps.items() if len(ids) > 10 and ky < _chip_cy for i in ids}
            _input_bumps = [b for b in _bumps if id(b) in _input_ids]
            _input_top_y = max(b["cy"] + b["h"] / 2 for b in _input_bumps) if _input_bumps else _chip_cy
            _output_ids = {i for ky, ids in _y_grps.items() if len(ids) > 10 and ky > _input_top_y for i in ids}

            for _b in _bumps:
                if id(_b) in _input_ids: _b["_cat"] = "INPUT"
                elif id(_b) in _output_ids: _b["_cat"] = "OUTPUT"
                else: _b["_cat"] = "DUMMY"

            # 統計 Bump Type 種數（W×H 組合）
            from collections import Counter as _Counter
            _wh_counts = _Counter((round(b["w"], 1), round(b["h"], 1)) for b in _bumps)
            _bump_type_count = len(_wh_counts)

            bump_type_str = f"{_bump_type_count} 種"
        except Exception as _e:
            print(f"[WARN] Bump type calculation failed: {_e}", file=_sys.stderr)
            bump_type_str = "—"

        display_name = (_uploaded_gds_name + Path(path).suffix) if _uploaded_gds_name else info["filename"]
        return jsonify({
            "rows": table_rows,
            "mapping_loaded": mapping_loaded,
            "mapping_path":   str(mapping_json) if mapping_loaded else "",
            "info": {
                "file":   display_name,
                "bump_type": bump_type_str,
                "layers": info["layers"],
                "unit":   "1 μm",
                "top":    info["tops"][0] if info["tops"] else "—",
            }
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ── Debug ─────────────────────────────────────────────────────────────────────
@app.route("/api/debug")
def api_debug():
    return jsonify({
        "LAYER_OVERRIDES":      LAYER_OVERRIDES,
        "LAYER_TEXT_OVERRIDES": LAYER_TEXT_OVERRIDES,
        "file": __file__,
    })

# ── 儲存 Layer Mapping JSON ────────────────────────────────────────────────────
@app.route("/api/load_mapping", methods=["POST"])
def api_load_mapping():
    """加載 Mapping JSON 文件並返回其中的層資訊（優先 output，再試 input）"""
    body = request.json or {}
    gds_path = body.get("file_path", "").strip()

    if not gds_path:
        return jsonify({"error": "GDS 路徑無效"}), 400

    try:
        stem = _uploaded_gds_name if _uploaded_gds_name else (Path(gds_path).stem if gds_path else "design")
        pname = _project_name(stem)

        # 優先從 output 目錄尋找，再試 input 目錄
        mapping_paths = [
            _HERE / "output" / f"{pname}.layer_mapping.json",
            _HERE / "input" / f"{pname}.layer_mapping.json",
        ]

        mapping_path = None
        for path in mapping_paths:
            if path.exists():
                mapping_path = path
                break

        if not mapping_path:
            return jsonify({"error": f"Mapping 文件不存在（已查找 output 和 input 目錄）"}), 404

        with open(mapping_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # 返回 all_layers 中的所有層資訊（包含 purpose）+ bump_type_stats
        all_layers = data.get("all_layers", [])
        bump_type_stats = data.get("bump_type_stats", [])
        return jsonify({
            "path": str(mapping_path),
            "rows": all_layers,
            "bump_type_stats": bump_type_stats
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/save_mapping", methods=["POST"])
def api_save_mapping():
    body = request.json or {}
    gds_path = body.get("file_path", "").strip()
    out_dir  = body.get("out_dir", "").strip()
    rows     = body.get("rows", [])
    bump_type_stats = body.get("bump_type_stats", [])

    if not rows:
        return jsonify({"error": "無 layer 資料"}), 400

    try:
        if not out_dir:
            out_dir = str(_HERE / "output")
        Path(out_dir).mkdir(parents=True, exist_ok=True)

        stem     = _uploaded_gds_name if _uploaded_gds_name else (Path(gds_path).stem if gds_path else "design")
        pname    = _project_name(stem)
        out_path = str(Path(out_dir) / f"{pname}.layer_mapping.json")

        mapping = {p: [] for p in PURPOSES}
        for r in rows:
            if r.get("purpose", "unknown") != "unknown":
                mapping[r["purpose"]].append([r["layer"], r["dt"]])

        data = {
            "mapping":        {k: v for k, v in mapping.items() if v},
            "all_layers":     rows,
            "bump_type_stats": bump_type_stats,  # 保存 Bump Type 分類
        }
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        return jsonify({"path": out_path})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ── GDS 轉圖片 ────────────────────────────────────────────────────────────────
@app.route("/api/run_image", methods=["POST"])
def api_run_image():
    body     = request.json or {}
    gds_path = body.get("file_path", "").strip()
    out_dir  = body.get("out_dir", "").strip()

    if not gds_path or not Path(gds_path).exists():
        return jsonify({"error": "GDS 路徑無效"}), 400

    try:
        if not out_dir:
            out_dir = str(_HERE / "output")
        Path(out_dir).mkdir(parents=True, exist_ok=True)
        # 原始檔名：優先用上傳時記錄的，否則用暫存路徑的 stem
        title = _uploaded_gds_name if _uploaded_gds_name else Path(gds_path).stem
        out_path = str(Path(out_dir) / f"{title}.png")

        _gds_to_image(gds_path, out_path, dpi=300, title=title)
        url = f"http://127.0.0.1:5000/output/{Path(out_path).name}"
        return jsonify({"path": out_path, "url": url, "message": f"已輸出：{out_path}"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def _gds_to_image(gds_path, out_path, dpi=300, title=None):
    """渲染 GDS top cell 成 PNG 圖片（白底黑線，無文字標注）"""
    import gdstk, matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import numpy as np

    lib   = gdstk.read_gds(gds_path)
    scale = lib.unit / 1e-6
    tops  = lib.top_level()
    if not tops:
        raise ValueError("No top-level cell")

    # Flatten top cell → 取得所有 polygon 的絕對座標（套用 reference 位移/旋轉）
    top_flat = tops[0].copy("_render_flat", deep_copy=True)
    top_flat.flatten()
    flat_polys = top_flat.polygons

    if not flat_polys:
        raise ValueError("GDS 無可繪製的 polygon")

    # 計算整體 bounding box（使用絕對座標）
    all_pts_np = np.vstack([p.points for p in flat_polys]) * scale
    xmin_view = float(all_pts_np[:,0].min())
    xmax_view = float(all_pts_np[:,0].max())
    ymin_view = float(all_pts_np[:,1].min())
    ymax_view = float(all_pts_np[:,1].max())

    W = xmax_view - xmin_view
    H = ymax_view - ymin_view
    aspect = W / H if H > 0 else 4.0

    # 緊密裁切：上方留尺寸標注空間，左側留高度標注，收緊垂直留白
    pad_x = W * 0.05
    pad_y = H * 0.22       # 比例收緊，讓晶片佔滿圖面高度
    xmin_lim = xmin_view - pad_x;  xmax_lim = xmax_view + pad_x
    ymin_lim = ymin_view - pad_y;  ymax_lim = ymax_view + pad_y

    # 輸出像素：寬 10000px（FHD 最低 1920px），高度依晶片比例計算（最低 1080px）
    TARGET_W_PX = max(10000, 1920)
    TARGET_H_PX = max(1080, int(TARGET_W_PX / aspect))
    fig_w = TARGET_W_PX / dpi
    fig_h = TARGET_H_PX / dpi

    fig, ax = plt.subplots(figsize=(fig_w, fig_h), facecolor="white")
    ax.set_facecolor("white")
    ax.tick_params(colors="#333333", labelsize=7)
    for sp in ax.spines.values():
        sp.set_edgecolor("#aaaaaa")

    # 判斷 bump layer（從 LAYER_OVERRIDES 取得）
    bump_layers = {lyr for lyr, purpose in LAYER_OVERRIDES.items() if purpose == "bump"}

    # 繪製所有 polygon
    for poly in flat_polys:
        pts = np.array(poly.points) * scale
        is_bump = poly.layer in bump_layers
        # chip_size：面積超大 → 只畫邊框，無填充
        _trapz = getattr(np, "trapezoid", None) or getattr(np, "trapz")
        area_um2 = abs(float(_trapz(pts[:,1], pts[:,0])))
        is_chip  = (area_um2 > 1e6)
        if is_chip:
            fc, ec, lw, zo = "none", "#333333", 1.2, 1
        elif is_bump:
            fc, ec, lw, zo = "#ffe066", "black", 0.8, 2
        else:
            fc, ec, lw, zo = "white", "black", 0.6, 1
        patch = plt.Polygon(pts, closed=True,
                            facecolor=fc, edgecolor=ec,
                            linewidth=lw, zorder=zo)
        ax.add_patch(patch)


    # 尺寸標注（以主體 bbox 為基準）
    arrow_kw = dict(arrowstyle="<->", color="#222222", lw=1.2)
    y_top = ymax_view + H * 0.18
    ax.annotate("", xy=(xmax_view, y_top), xytext=(xmin_view, y_top),
                arrowprops=arrow_kw)
    ax.text((xmin_view+xmax_view)/2, y_top + H*0.06,
            f"{W:.1f} μm (Width)",
            ha="center", va="bottom", fontsize=8, color="#222222",
            fontweight="bold")

    x_left = xmin_view - W * 0.04
    ax.annotate("", xy=(x_left, ymax_view), xytext=(x_left, ymin_view),
                arrowprops=arrow_kw)
    ax.text(x_left - W*0.005, (ymin_view+ymax_view)/2,
            f"{H:.1f} μm\n(Height)",
            ha="right", va="center", fontsize=8, color="#222222",
            fontweight="bold", rotation=90)

    # IC chip 中心準星（紅色）
    cx = (xmin_view + xmax_view) / 2
    cy = (ymin_view + ymax_view) / 2
    arm_x = W * 0.015
    arm_y = H * 0.25
    ax.plot([cx - arm_x, cx + arm_x], [cy, cy], color="red", lw=1.5, zorder=10)
    ax.plot([cx, cx], [cy - arm_y, cy + arm_y], color="red", lw=1.5, zorder=10)
    ax.plot(cx, cy, "s", color="red", markersize=4, zorder=11)
    ax.text(cx + arm_x * 0.3, cy + arm_y * 1.1, "(0, 0)",
            fontsize=7, color="red", ha="center", va="bottom")

    ax.set_xlabel("X (μm)", color="#333333", fontsize=8)
    ax.set_ylabel("Y (μm)", color="#333333", fontsize=8)
    ax.set_title(title or Path(gds_path).stem, color="#2c3e50", fontsize=11, fontweight="bold")
    ax.set_xlim(xmin_lim, xmax_lim)
    ax.set_ylim(ymin_lim, ymax_lim)
    ax.set_aspect("equal", adjustable="datalim")
    ax.grid(True, linestyle="--", linewidth=0.3, color="#eeeeee", alpha=0.8)

    fig.savefig(out_path, dpi=dpi, bbox_inches="tight", facecolor="white")
    plt.close(fig)


# ── Bump / Amark 元件分析 ─────────────────────────────────────────────────────
@app.route("/api/run_elements", methods=["POST"])
def api_run_elements():
    """Bump/Amark 分析：使用與 PAD Excel 相同的分析邏輯（test_export_xlsx）"""
    body     = request.json or {}
    gds_path = body.get("file_path", "").strip()
    out_dir  = body.get("out_dir", "").strip()

    if not gds_path or not Path(gds_path).exists():
        return jsonify({"error": "GDS 路徑無效"}), 400

    try:
        import importlib, sys as _sys, numpy as _np
        if not out_dir:
            out_dir = str(_HERE / "output")
        stem  = _uploaded_gds_name if _uploaded_gds_name else Path(gds_path).stem
        pname = _project_name(stem)
        mj    = str(Path(out_dir) / f"{pname}.layer_mapping.json")

        # 載入 test_export_xlsx 模組（與 Excel 輸出共用同一套邏輯）
        if "test_export_xlsx" in _sys.modules:
            _mod = importlib.reload(_sys.modules["test_export_xlsx"])
        else:
            _mod = importlib.import_module("test_export_xlsx")

        # 套用 layer mapping
        if Path(mj).exists():
            _mdata = json.loads(Path(mj).read_text(encoding="utf-8"))
            _m = _mdata.get("mapping", {})
            def _first(k): lst = _m.get(k, []); return tuple(lst[0]) if lst else None
            if _first("bump"):             _mod.BUMP_LAYER      = _first("bump")
            if _first("amark"):            _mod.AMARK_LAYER     = _first("amark")
            if _first("bump_text_number"): _mod.TEXT_NUM_LAYER  = _first("bump_text_number")
            if _first("bump_text_name"):   _mod.TEXT_NAME_LAYER = _first("bump_text_name")

        top, scale = _mod.load_flat(gds_path)
        chip_w, chip_h = _mod.get_chip_size(top, scale)
        bumps  = _mod.get_bumps(top, scale)
        amarks = _mod.get_amarks(top, scale)

        # 空間分類（密集單排 + 晶片下半 → INPUT，上半 → OUTPUT，其餘 → DUMMY）
        _chip_bbox = _mod.get_chip_bbox(top, scale)
        _chip_cy   = (_chip_bbox[1] + _chip_bbox[3]) / 2 if _chip_bbox else 0

        _y_grps: dict = {}
        for _b in bumps:
            _ky = round(_b["cy"] * 2) / 2
            _y_grps.setdefault(_ky, []).append(id(_b))

        _input_ids   = {i for ky, ids in _y_grps.items() if len(ids) > 10 and ky < _chip_cy for i in ids}
        _input_bumps = [b for b in bumps if id(b) in _input_ids]
        _input_top_y = max(b["cy"] + b["h"] / 2 for b in _input_bumps) if _input_bumps else _chip_cy
        _output_ids  = {i for ky, ids in _y_grps.items() if len(ids) > 10 and ky > _input_top_y for i in ids}

        for _b in bumps:
            if id(_b) in _input_ids:   _b["_cat"] = "INPUT"
            elif id(_b) in _output_ids: _b["_cat"] = "OUTPUT"
            else:                       _b["_cat"] = "DUMMY"

        def _classify(b):
            return b.get("_cat", "DUMMY")

        input_n  = sum(1 for b in bumps if _classify(b) == "INPUT")
        output_n = sum(1 for b in bumps if _classify(b) == "OUTPUT")
        dummy_n  = sum(1 for b in bumps if _classify(b) == "DUMMY")

        # Bump Type 統計（每種 W×H 的數量與分類）
        from collections import Counter as _Counter
        _wh_counts = _Counter((b["w"], b["h"]) for b in bumps)

        # 統計每個 W×H 的各分類數量，取最多的分類
        _wh_cat_counts = {}
        for _b in bumps:
            _k = (_b["w"], _b["h"])
            _c = _b.get("_cat", "DUMMY")
            if _k not in _wh_cat_counts:
                _wh_cat_counts[_k] = _Counter()
            _wh_cat_counts[_k][_c] += 1

        _wh_cat = {}
        for _k, _c_cnt in _wh_cat_counts.items():
            # 取該 W×H 中數量最多的分類
            _wh_cat[_k] = _c_cnt.most_common(1)[0][0] if _c_cnt else "DUMMY"

        type_stats = sorted([
            {"w": bw, "h": bh, "count": cnt, "cat": _wh_cat.get((bw, bh), "DUMMY")}
            for (bw, bh), cnt in _wh_counts.items()
        ], key=lambda x: -x["count"])

        # 載入已儲存的 Bump Type 分類（如果存在）
        bump_type_stats_saved = []
        if Path(mj).exists():
            try:
                _mdata = json.loads(Path(mj).read_text(encoding="utf-8"))
                bump_type_stats_saved = _mdata.get("bump_type_stats", [])
                if bump_type_stats_saved:
                    # 建立 (w, h) → cat 對應表（使用四捨五入到 4 位小數以容納浮點精度差異）
                    _saved_lookup = {}
                    for _b in bump_type_stats_saved:
                        _w = round(float(_b["w"]), 4)
                        _h = round(float(_b["h"]), 4)
                        _saved_lookup[(_w, _h)] = _b["cat"]
                    # 覆蓋 type_stats 的 cat 欄位
                    for _ts in type_stats:
                        _key = (round(_ts["w"], 4), round(_ts["h"], 4))
                        if _key in _saved_lookup:
                            _ts["cat"] = _saved_lookup[_key]
            except Exception:
                pass

        msg = (
            f"Chip Size: {chip_w} × {chip_h} μm\n"
            f"Bumps: {len(bumps)}  (INPUT {input_n} / OUTPUT {output_n} / DUMMY {dummy_n})\n"
            f"Amarks: {len(amarks)}\n"
        )
        for i, (ax, ay) in enumerate(amarks, 1):
            msg += f"  A{i}: ({ax}, {ay}) μm\n"

        return jsonify({"message": msg, "bumps": len(bumps),
                        "input": input_n, "output": output_n, "dummy": dummy_n,
                        "amarks": amarks, "chip_w": chip_w, "chip_h": chip_h,
                        "type_stats": type_stats,
                        "bump_type_stats": bump_type_stats_saved})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def _versioned_path(base_path: str) -> str:
    """若檔案已存在，自動附加 _v1, _v2, ... 直到找到未使用的路徑"""
    p = Path(base_path)
    if not p.exists():
        return base_path
    parent, stem, suffix = p.parent, p.stem, p.suffix
    v = 1
    while True:
        candidate = parent / f"{stem}_v{v}{suffix}"
        if not candidate.exists():
            return str(candidate)
        v += 1

# ── Excel PAD 資料輸出 ────────────────────────────────────────────────────────
def _build_xlsx(gds_path: str, out_dir: str, original_stem: str = None,
                mapping_json: str = None, overrides: dict = None, window_um: int = 3500) -> str:
    """委派給 test_export_xlsx.run()（定案版佈局）

    Args:
        window_um: 簡圖各區域寬度（µm），預設 3500，A2 Center 固定 ±1750
    """
    import importlib, sys as _sys
    stem         = original_stem if original_stem else Path(gds_path).stem
    pname        = _project_name(stem)
    out_path     = _versioned_path(str(Path(out_dir) / f"{stem}.xlsx"))
    if mapping_json is None:
        mapping_json = str(Path(out_dir) / f"{pname}.layer_mapping.json")
    if "test_export_xlsx" in _sys.modules:
        _mod = importlib.reload(_sys.modules["test_export_xlsx"])
    else:
        _mod = importlib.import_module("test_export_xlsx")
    _mod.run(gds_path, out_path, mapping_json=mapping_json,
             original_stem=stem, overrides=overrides or {}, window_um=window_um)
    return out_path

def _build_xlsx_UNUSED(gds_path: str, out_dir: str, original_stem: str = None, mapping_json: str = None) -> str:
    """從 GDS 生成 2-sheet Excel（PAD清單 + INFORMATION），回傳輸出路徑"""
    import io as _io, re, json as _json
    import numpy as np
    import gdstk
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    stem = original_stem if original_stem else Path(gds_path).stem

    # ── 讀取 layer mapping（優先用 JSON，否則用預設值）───────────────────────
    if mapping_json is None:
        mapping_json = str(Path(out_dir) / f"{stem}.layer_mapping.json")
    bump_layers      = set()
    amark_layers     = set()
    text_num_layers  = set()
    text_name_layers = set()
    chip_size_layers = set()
    if Path(mapping_json).exists():
        _mdata = _json.loads(Path(mapping_json).read_text(encoding="utf-8"))
        _m = _mdata.get("mapping", {})
        for _ld in _m.get("bump",             []): bump_layers.add(tuple(_ld))
        for _ld in _m.get("amark",            []): amark_layers.add(tuple(_ld))
        for _ld in _m.get("bump_text_number", []): text_num_layers.add(tuple(_ld))
        for _ld in _m.get("bump_text_name",   []): text_name_layers.add(tuple(_ld))
        for _ld in _m.get("chip_size",        []): chip_size_layers.add(tuple(_ld))
    # fallback
    if not bump_layers:      bump_layers      = {(225, 0)}
    if not amark_layers:     amark_layers     = {(205, 224)}
    if not text_num_layers:  text_num_layers  = {(225, 0)}
    if not text_name_layers: text_name_layers = {(211, 224)}
    BUMP_LAYER  = next(iter(bump_layers))   # 主 bump layer（繪圖用）
    AMARK_LAYER = next(iter(amark_layers))  # 主 amark layer（繪圖用）

    # ── GDS 載入 (flatten) ───────────────────────────────────────────────────
    lib   = gdstk.read_gds(str(gds_path))
    scale = lib.unit / 1e-6
    top   = lib.top_level()[0].copy("_flat", deep_copy=True)
    top.flatten()

    # ── 晶片尺寸（優先用 chip_size layer，否則取最大矩形）────────────────────
    chip_w = chip_h = None
    best_area = 0
    for poly in top.polygons:
        if len(poly.points) == 4:
            key = (poly.layer, poly.datatype)
            if chip_size_layers and key not in chip_size_layers:
                continue
            pts = np.array(poly.points) * scale
            w = float(pts[:,0].max() - pts[:,0].min())
            h = float(pts[:,1].max() - pts[:,1].min())
            if w * h > best_area:
                best_area = w * h
                chip_w, chip_h = round(w, 4), round(h, 4)
    # 若 chip_size layer 找不到，退回全圖最大矩形
    if chip_w is None:
        for poly in top.polygons:
            if len(poly.points) == 4:
                pts = np.array(poly.points) * scale
                w = float(pts[:,0].max() - pts[:,0].min())
                h = float(pts[:,1].max() - pts[:,1].min())
                if w * h > best_area:
                    best_area = w * h
                    chip_w, chip_h = round(w, 4), round(h, 4)

    # ── Amark 位置 ───────────────────────────────────────────────────────────
    seen_am, amarks = set(), []
    for poly in top.polygons:
        if (poly.layer, poly.datatype) in amark_layers and len(poly.points) == 4:
            pts = np.array(poly.points) * scale
            cx = round(float(pts[:,0].mean()), 3)
            cy = round(float(pts[:,1].mean()), 3)
            if (cx, cy) not in seen_am:
                seen_am.add((cx, cy)); amarks.append((cx, cy))
    amarks.sort()

    # ── Bump 清單 ────────────────────────────────────────────────────────────
    bumps = []
    for poly in top.polygons:
        if (poly.layer, poly.datatype) in bump_layers:
            pts = np.array(poly.points) * scale
            w = float(pts[:,0].max()-pts[:,0].min())
            h = float(pts[:,1].max()-pts[:,1].min())
            if w > 5000: continue
            bumps.append({"cx": round(float(pts[:,0].mean()),4),
                          "cy": round(float(pts[:,1].mean()),4),
                          "w": round(w,4), "h": round(h,4),
                          "number": "", "name": ""})
    num_lbs, name_lbs = [], []
    for lb in top.labels:
        lx = round(lb.origin[0]*scale, 4)
        ly = round(lb.origin[1]*scale, 4)
        if (lb.layer, lb.texttype) in text_num_layers and "NO" in lb.text.upper():
            num_lbs.append((lx, ly, lb.text))
        if (lb.layer, lb.texttype) in text_name_layers:
            name_lbs.append((lx, ly, lb.text))
    if bumps:
        bxy = np.array([[b["cx"],b["cy"]] for b in bumps])
        for lx,ly,txt in num_lbs:
            bumps[int(np.argmin(np.linalg.norm(bxy-[lx,ly],axis=1)))]["number"] = txt
        for lx,ly,txt in name_lbs:
            bumps[int(np.argmin(np.linalg.norm(bxy-[lx,ly],axis=1)))]["name"] = txt
    def _no_key(b):
        s = b["number"].upper().replace("NO","").strip()
        return int(s) if s.isdigit() else 999999
    bumps.sort(key=_no_key)

    # ── 空間分類：INPUT上緣以上，密集行(同Y>10個)→OUTPUT，稀疏→DUMMY ────────
    _input_bumps = [b for b in bumps if round(b["w"]) == 24]
    _input_top_y = max(b["cy"] + b["h"] / 2 for b in _input_bumps) if _input_bumps else 0
    # 對 INPUT 上方的 bump 按 Y 層分群（容差 ±10 μm）
    _output_set: set = set()
    if _input_bumps:
        _above = [b for b in bumps if b["cy"] > _input_top_y]
        _y_groups: dict = {}
        for _b in _above:
            _ky = round(_b["cy"] * 2) / 2
            _y_groups.setdefault(_ky, []).append(id(_b))
        for _ky, _ids in _y_groups.items():
            if len(_ids) > 10:   # 密集行 → OUTPUT
                _output_set.update(_ids)

    def classify(b):
        if round(b["w"]) == 24: return "INPUT"
        if id(b) in _output_set:  return "OUTPUT"
        return "DUMMY"

    # ── 樣式 helpers ─────────────────────────────────────────────────────────
    def _bdr(c="CCCCCC"):
        s = Side(style="thin", color=c)
        return Border(left=s, right=s, top=s, bottom=s)
    def hdr(ws, r, c, val, bg="1F4E79", fg="FFFFFF"):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font      = Font(bold=True, color=fg, size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _bdr("999999")
    def dat(ws, r, c, val, align="center", fmt=None, bold=False, bg=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = _bdr()
        if fmt:  cell.number_format = fmt
        if bold: cell.font = Font(bold=True, size=10)
        if bg:   cell.fill = PatternFill("solid", fgColor=bg)

    # ── Sheet 1：PAD 清單 ─────────────────────────────────────────────────────
    stem = original_stem if original_stem else Path(gds_path).stem
    sheet_name = stem[:31]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = sheet_name

    ws1.cell(row=1, column=1, value="Chip_Size").font = Font(bold=True, size=11)
    ws1.cell(row=1, column=2, value="include SR+SL")
    ws1.cell(row=1, column=3, value=f"{chip_w} * {chip_h}" if chip_w else "—")
    ws1.cell(row=2, column=1, value="PAD Location").font = Font(bold=True, size=11)
    for c, h in enumerate(["No.","Pad Name","X-axis(um)","Y-axis(um)","W(um)","H(um)","Area"], 1):
        hdr(ws1, 3, c, h)

    input_area  = sum(round(b["w"])*round(b["h"]) for b in bumps if classify(b)=="INPUT")
    output_area = sum(round(b["w"])*round(b["h"]) for b in bumps if classify(b)=="OUTPUT")
    dummy_area  = sum(round(b["w"])*round(b["h"]) for b in bumps if classify(b)=="DUMMY")
    for i,(lbl,val) in enumerate([("INPUT Bump area",input_area),("OUTPUT Bump area",output_area),
                                   ("DUMMY Bump area",dummy_area),("Total Bump area",input_area+output_area+dummy_area)]):
        r = 3+i
        cl = ws1.cell(row=r, column=9, value=lbl)
        cl.font=Font(bold=True,color="FFFFFF",size=10); cl.fill=PatternFill("solid",fgColor="002060")
        cl.alignment=Alignment(horizontal="right",vertical="center"); cl.border=_bdr("999999")
        cv = ws1.cell(row=r, column=10, value=val)
        cv.number_format="#,##0.00"; cv.alignment=Alignment(horizontal="center",vertical="center")
        cv.border=_bdr("999999"); cv.fill=PatternFill("solid",fgColor="BDD7EE")
        ws1.cell(row=r, column=11, value="um^2").alignment=Alignment(horizontal="left")

    CAT_BG = {"INPUT":"EBF3FB","OUTPUT":"E2EFDA","DUMMY":"FFF2CC"}
    for i, b in enumerate(bumps, 4):
        cat = classify(b); bg = CAT_BG.get(cat,"FFFFFF")
        dat(ws1,i,1,b["number"],bg=bg); dat(ws1,i,2,b["name"],align="left",bg=bg)
        dat(ws1,i,3,str(b["cx"]),bg=bg); dat(ws1,i,4,str(b["cy"]),bg=bg)
        dat(ws1,i,5,str(b["w"]),bg=bg);  dat(ws1,i,6,str(b["h"]),bg=bg)
        dat(ws1,i,7,round(b["w"])*round(b["h"]),bg=bg)
    for col,w in {"A":12,"B":22,"C":14,"D":14,"E":8,"F":8,"G":10,"I":18,"J":16,"K":8}.items():
        ws1.column_dimensions[col].width = w
    ws1.row_dimensions[3].height = 22

    # ── 圖片生成 helpers ─────────────────────────────────────────────────────
    _trapz = getattr(np, "trapezoid", None) or getattr(np, "trapz")

    def _chip_img() -> bytes:
        all_pts = np.vstack([np.array(p.points)*scale for p in top.polygons if len(p.points)>=3])
        xmn=float(all_pts[:,0].min()); xmx=float(all_pts[:,0].max())
        ymn=float(all_pts[:,1].min()); ymx=float(all_pts[:,1].max())
        W=xmx-xmn; H=ymx-ymn
        aspect=W/H if H>0 else 1
        fig,ax=plt.subplots(figsize=(22,max(3,22/aspect)),dpi=200)
        ax.set_facecolor("white"); fig.patch.set_facecolor("white")
        for poly in top.polygons:
            pts=np.array(poly.points)*scale
            key=(poly.layer,poly.datatype)
            pw=float(pts[:,0].max()-pts[:,0].min())
            area=abs(float(_trapz(pts[:,1],pts[:,0])))
            is_chip=(area>1e7); is_bump=(key in bump_layers and pw<5000)
            if is_chip:   fc,ec,lw="none","#222222",1.5
            elif is_bump: fc,ec,lw="#ffe066","black",0.3
            else:         fc,ec,lw="white","#555555",0.5
            ax.add_patch(plt.Polygon(pts,closed=True,facecolor=fc,edgecolor=ec,linewidth=lw,zorder=2 if is_bump else 1))
        pad=W*0.02
        ax.set_xlim(xmn-pad,xmx+pad); ax.set_ylim(ymn-H*0.25,ymx+H*0.12)
        ax.set_aspect("equal"); ax.axis("off")
        arr_y=ymx+H*0.06
        ax.annotate("",xy=(xmx,arr_y),xytext=(xmn,arr_y),arrowprops=dict(arrowstyle="<->",color="red",lw=1.2))
        ax.text((xmn+xmx)/2,arr_y+H*0.03,f"{chip_w}um (SR + SL)",ha="center",va="bottom",fontsize=9,color="red",fontweight="bold")
        arr_x=xmn-W*0.015
        ax.annotate("",xy=(arr_x,ymx),xytext=(arr_x,ymn),arrowprops=dict(arrowstyle="<->",color="red",lw=1.2))
        ax.text(arr_x-W*0.005,(ymn+ymx)/2,f"{chip_h}um\n(SR + SL)",ha="right",va="center",fontsize=9,color="red",fontweight="bold",rotation=90)
        cx=(xmn+xmx)/2; cy=(ymn+ymx)/2
        ax.plot([cx-W*0.008,cx+W*0.008],[cy,cy],color="red",lw=1.2,zorder=10)
        ax.plot([cx,cx],[cy-H*0.15,cy+H*0.15],color="red",lw=1.2,zorder=10)
        ax.text(cx+W*0.006,cy+H*0.18,"(0,0)",fontsize=8,color="red",ha="left",va="bottom")
        ax.text(cx+W*0.025,cy,"X",fontsize=10,color="red",ha="left",va="center",fontweight="bold")
        ax.text(cx,cy+H*0.18,"Y",fontsize=10,color="red",ha="center",va="bottom",fontweight="bold")
        buf=_io.BytesIO()
        fig.savefig(buf,format="png",dpi=200,bbox_inches="tight",facecolor="white")
        plt.close(fig); buf.seek(0); return buf.read()

    def _amark_img(ax_um, ay_um, label) -> bytes:
        raw=[]
        for poly in top.polygons:
            pts=np.array(poly.points)*scale
            cx=float(pts[:,0].mean()); cy=float(pts[:,1].mean())
            pw=float(pts[:,0].max()-pts[:,0].min())
            if abs(cx-ax_um)<200 and abs(cy-ay_um)<200 and pw<200:
                raw.append((poly.layer,poly.datatype,len(poly.points),pts-np.array([ax_um,ay_um])))
        seen2,unique=[],[]
        for layer,dtype,n,pts in raw:
            key=(layer,dtype,n,tuple(map(tuple,np.round(pts,1).tolist())))
            if key not in seen2: seen2.append(key); unique.append((layer,dtype,n,pts))
        outer=None; inner=[]
        for layer,dtype,n,pts in unique:
            if (layer,dtype)==AMARK_LAYER: outer=pts
            elif (layer,dtype)!=BUMP_LAYER: inner.append((n,pts))
        if outer is None: return b""
        O=50.0; MARGIN=24; is_cross=any(n>=10 for n,_ in inner)
        fig,ap=plt.subplots(figsize=(9.6,9.6),dpi=200)
        ap.set_facecolor("white"); fig.patch.set_facecolor("white")
        ap.set_aspect("equal"); ap.axis("off")
        ap.set_xlim(-O-MARGIN,O+MARGIN); ap.set_ylim(-O-MARGIN-4,O+MARGIN+6)
        ap.add_patch(plt.Polygon(outer,closed=True,facecolor="none",edgecolor="black",linewidth=1.5,zorder=2))
        for n,pts in inner:
            ap.add_patch(plt.Polygon(pts,closed=True,facecolor="none",edgecolor="black",linewidth=1.0,zorder=3))
        ap.text(-O-MARGIN+1,O+MARGIN+4,label,ha='left',va='top',fontsize=12,fontweight='bold',color='red')
        AP=dict(arrowstyle='<->',color='black',lw=0.8,mutation_scale=9)
        def dim_h(x1,x2,y_base,y_line,text,fs=8.5):
            if abs(y_line-y_base)>0.3:
                ap.plot([x1,x1],[y_base,y_line],'k-',lw=0.5,zorder=5)
                ap.plot([x2,x2],[y_base,y_line],'k-',lw=0.5,zorder=5)
            ap.annotate('',xy=(x2,y_line),xytext=(x1,y_line),arrowprops=AP)
            sign=1 if y_line>=y_base else -1
            ap.text((x1+x2)/2,y_line+sign*2.5,text,ha='center',va='bottom' if sign>0 else 'top',fontsize=fs)
        def dim_v(y1,y2,x_base,x_line,text,fs=8.5):
            if abs(x_line-x_base)>0.3:
                ap.plot([x_base,x_line],[y1,y1],'k-',lw=0.5,zorder=5)
                ap.plot([x_base,x_line],[y2,y2],'k-',lw=0.5,zorder=5)
            ap.annotate('',xy=(x_line,y2),xytext=(x_line,y1),arrowprops=AP)
            sign=1 if x_line>=x_base else -1
            ap.text(x_line+sign*2.5,(y1+y2)/2,text,ha='left' if sign>0 else 'right',va='center',fontsize=fs,rotation=90)
        if is_cross:
            A,B=12.5,37.5
            dim_h(-O,O,O,O+15,'100'); dim_v(-O,O,-O,-O-15,'100')
            dim_v(B,O,A,A+10,'12.5',fs=7.5); dim_h(-O,-B,-A,-A-10,'12.5',fs=7.5)
            dim_h(-A,A,B,B+8,'25',fs=7.5); dim_v(-A,A,B,B+8,'25',fs=7.5)
            dim_v(A,B,A,A+8,'25',fs=7.5);  dim_h(A,B,A,A+8,'25',fs=7.5)
        else:
            C,T=12.5,37.5
            dim_h(-O,O,O,O+15,'100'); dim_v(-O,O,-O,-O-15,'100')
            BOT=-O-11
            dim_h(-O,-C,-O,BOT,'37.5',fs=7.5); dim_h(-C,C,-O,BOT,'25',fs=7.5); dim_h(C,O,-O,BOT,'37.5',fs=7.5)
            RGT=O+11
            dim_v(-O,-C,O,RGT,'37.5',fs=7.5); dim_v(-C,C,O,RGT,'25',fs=7.5); dim_v(C,O,O,RGT,'37.5',fs=7.5)
            dim_h(C,T,C,C+9,'25',fs=7.5); dim_v(C,T,C,C+9,'25',fs=7.5)
        buf=_io.BytesIO()
        fig.savefig(buf,format="png",dpi=200,bbox_inches="tight",facecolor="white")
        plt.close(fig); buf.seek(0); return buf.read()

    def _embed(ws, png_bytes, anchor, width_px, height_px):
        if png_bytes:
            img=XLImage(_io.BytesIO(png_bytes)); img.width=width_px; img.height=height_px
            img.anchor=anchor; ws.add_image(img)

    # ── Sheet 2：INFORMATION ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("INFORMATION")
    CHIP_ROWS = 22
    _embed(ws2, _chip_img(), "A1", 1100, 220)
    for ri in range(1, CHIP_ROWS+1):
        ws2.row_dimensions[ri].height = 10

    def no_range(nos):
        return f"{nos[0]}~{nos[-1]}" if nos else ""

    input_t  = defaultdict(lambda: {"count":0,"nos":[]})
    output_t = defaultdict(lambda: {"count":0,"nos":[]})
    dummy_t  = defaultdict(lambda: {"count":0,"nos":[]})
    for b in bumps:
        cat=classify(b); key=(round(b["w"]),round(b["h"])); no=b["number"]
        if cat=="INPUT":   input_t[key]["count"]+=1; input_t[key]["nos"].append(no)
        elif cat=="OUTPUT":output_t[key]["count"]+=1;output_t[key]["nos"].append(no)
        else:              dummy_t[key]["count"]+=1; dummy_t[key]["nos"].append(no)

    r = CHIP_ROWS + 1
    ws2.cell(row=r,column=5,value="Condition-1 (+/- 3um temp. compensation)  ").font=Font(bold=True)
    ws2.cell(row=r,column=9,value="Unit=um")
    r += 1
    for col,lbl,bg in [(5,"INPUT PAD","1F4E79"),(9,"OUTPUT PAD","375623"),(13,"DUMMY PAD","7F6000"),(17,"AMARK","833C11")]:
        c=ws2.cell(row=r,column=col,value=lbl)
        c.font=Font(bold=True,color="FFFFFF",size=11); c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal="center",vertical="center")
    r += 1
    for col,lbl in [(5,"Symbol"),(6,"W(um)"),(7,"H(um)"),(8,"NO Range"),
                    (9,"Symbol"),(10,"W(um)"),(11,"H(um)"),(12,"NO Range"),
                    (13,"Symbol"),(14,"W(um)"),(15,"H(um)"),(16,"NO Range"),
                    (17,"Symbol"),(18,"W(um)"),(19,"H(um)")]:
        hdr(ws2,r,col,lbl,bg="2E75B6")
    r += 1
    max_rows=max(len(input_t),len(output_t),len(dummy_t),1)
    for i,(key,info) in enumerate(sorted(input_t.items())):
        dat(ws2,r+i,5,f"A{i+1}",bg="EBF3FB",bold=True); dat(ws2,r+i,6,key[0],bg="EBF3FB")
        dat(ws2,r+i,7,key[1],bg="EBF3FB");               dat(ws2,r+i,8,no_range(info["nos"]),align="left",bg="EBF3FB")
    for i,(key,info) in enumerate(sorted(output_t.items())):
        dat(ws2,r+i,9,f"B{i+1}",bg="E2EFDA",bold=True);  dat(ws2,r+i,10,key[0],bg="E2EFDA")
        dat(ws2,r+i,11,key[1],bg="E2EFDA");               dat(ws2,r+i,12,no_range(info["nos"]),align="left",bg="E2EFDA")
    for i,(key,info) in enumerate(sorted(dummy_t.items())):
        dat(ws2,r+i,13,f"C{i+1}",bg="FFF2CC",bold=True); dat(ws2,r+i,14,key[0],bg="FFF2CC")
        dat(ws2,r+i,15,key[1],bg="FFF2CC");               dat(ws2,r+i,16,no_range(info["nos"]),align="left",bg="FFF2CC")
    r2=r+max_rows+2
    ws2.cell(row=r2,column=4,value="AMARK").font=Font(bold=True,size=11)
    r2+=1
    for col,h in enumerate(["Symbol","X (um)","Y (um)"],4):
        hdr(ws2,r2,col,h,bg="833C11")
    r2+=1
    for i,(ax2,ay2) in enumerate(amarks,1):
        dat(ws2,r2,4,f"A{i}",bold=True); dat(ws2,r2,5,ax2,fmt="0.000"); dat(ws2,r2,6,ay2,fmt="0.000"); r2+=1
    img_row=r2+2
    for ri in range(img_row,img_row+15):
        ws2.row_dimensions[ri].height=14
    for i,(ax2,ay2) in enumerate(amarks):
        png=_amark_img(ax2,ay2,f"A{i+1}")
        _embed(ws2,png,f"{get_column_letter(1+i*5)}{img_row}",220,220)
    for col,w in [("A",12),("B",12),("C",12),("D",10),("E",12),("F",10),("G",10),("H",20),
                  ("I",10),("J",10),("K",10),("L",20),("M",10),("N",10),("O",10),("P",20),
                  ("Q",10),("R",10),("S",10)]:
        ws2.column_dimensions[col].width=w

    # ── 儲存 ─────────────────────────────────────────────────────────────────
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    out_path = str(Path(out_dir) / f"{stem}.xlsx")
    wb.save(out_path)
    return out_path


@app.route("/api/run_xlsx", methods=["POST"])
def api_run_xlsx():
    body      = request.json or {}
    gds_path  = body.get("file_path", _uploaded_gds_path).strip()
    out_dir   = body.get("out_dir", "").strip() or str(_HERE / "output")
    overrides = body.get("overrides", {})   # {"w,h": "INPUT/OUTPUT/DUMMY"}
    window_um = body.get("window_um", 3500)  # 簡圖區域寬度，預設 3500
    if not gds_path or not Path(gds_path).exists():
        return jsonify({"error": "GDS 路徑無效"}), 400
    try:
        stem  = _uploaded_gds_name or Path(gds_path).stem
        pname = _project_name(stem)
        mj    = str(Path(out_dir) / f"{pname}.layer_mapping.json")

        # 若 overrides 為空，從已儲存的 mapping JSON 加載 bump_type_stats
        if not overrides and Path(mj).exists():
            try:
                _mdata = json.loads(Path(mj).read_text(encoding="utf-8"))
                _bts = _mdata.get("bump_type_stats", [])
                if _bts:
                    overrides = {f"{b['w']},{b['h']}": b['cat'] for b in _bts}
            except Exception:
                pass

        out_path = _build_xlsx(gds_path, out_dir, original_stem=stem,
                               mapping_json=mj, overrides=overrides, window_um=window_um)
        return jsonify({"path": out_path, "message": f"已輸出：{out_path}"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ── Bump Rule 檢查 ────────────────────────────────────────────────────────────
@app.route("/api/check_bump_summary", methods=["POST"])
def api_check_bump_summary():
    """快速 Bump Layer 概要檢查"""
    body = request.json or {}
    gds_path = body.get("file_path", _uploaded_gds_path).strip()

    if not gds_path or not Path(gds_path).exists():
        return jsonify({"error": "GDS 路徑無效"}), 400

    try:
        from bump_checker import check_bump_rules_summary
        result = check_bump_rules_summary(gds_path)
        return jsonify(result)
    except ImportError as e:
        return jsonify({"error": f"模組導入失敗: {str(e)}"}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/run_bump_check", methods=["POST"])
def api_run_bump_check():
    """執行完整 Bump Rule 檢查（Perl 優先，無 Perl 時自動用 Python）"""
    body = request.json or {}
    gds_path = body.get("file_path", _uploaded_gds_path).strip()
    out_dir = body.get("out_dir", "").strip() or str(_HERE / "output")

    if not gds_path or not Path(gds_path).exists():
        return jsonify({"error": "GDS 路徑無效"}), 400

    try:
        from bump_checker import run_bump_check_perl, run_bump_check_python
        Path(out_dir).mkdir(parents=True, exist_ok=True)

        print(f"[DEBUG] Bump Check: gds_path={gds_path}, original_name={_uploaded_gds_name}")

        # 先嘗試 Perl 版本
        result = run_bump_check_perl(gds_path, out_dir, original_name=_uploaded_gds_name)

        # 如果 Perl 失敗（找不到 Perl），自動降級到 Python 版本
        if not result.get("success") and "Perl" in result.get("message", ""):
            print("[INFO] Perl not found, falling back to Python Bump Check...")
            result = run_bump_check_python(gds_path, out_dir, original_name=_uploaded_gds_name)
            result["note"] = "Using Python Bump Check (Perl not available)"

        return jsonify(result)
    except ImportError as e:
        return jsonify({"error": f"模組導入失敗: {str(e)}"}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ── 啟動 ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print("  GDS Reader Server  —  http://localhost:5000")
    print("=" * 55)
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
