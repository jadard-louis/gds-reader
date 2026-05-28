#!/usr/bin/env python3
"""
獨立測試：從 GDS 生成 2-sheet Excel（PAD_DATA / INFORMATION）
格式完全對照 JD9365TM_PAD_IN_OUT_SHIFT_7D5UM_20260210_Y1065D3.xlsx
"""
import sys
sys.path.insert(0, '.')

from pathlib import Path
from collections import defaultdict
import io
import struct
import numpy as np
from PIL import Image as _PilImg, ImageDraw as _PilDraw
import gdstk
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyArrowPatch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

GDS_PATH = "input/JD9365TM_PAD_IN_OUT_SHIFT_7D5UM_20260209_Y1065D3.gds"

def _next_out_path(base="output/test_export", ext=".xlsx"):
    """自動找下一個可用的遞增編號輸出路徑"""
    import re
    from pathlib import Path
    out_dir = Path(base).parent
    stem = Path(base).name
    existing = sorted(out_dir.glob(f"{stem}*.{ext.lstrip('.')}"))
    max_n = 0
    for p in existing:
        m = re.search(r"(\d+)$", p.stem)
        if m:
            max_n = max(max_n, int(m.group(1)))
    return str(out_dir / f"{stem}{max_n + 1}{ext}")

OUT_PATH = _next_out_path()

BUMP_LAYERS     = [(224, 0), (225, 0)]  # 強制對應：Layer 224/225 DType 0 → BUMP/PAD
CHIP_LAYER      = (190, 9)
AMARK_LAYER     = (205, 224)
TEXT_NUM_LAYERS = [(224, 0), (225, 0)]  # 強制對應：Layer 224/225 DType 0 → Bump 序號標注
TEXT_NAME_LAYER = (211, 224)

# ── GDS 讀取（flatten 絕對座標）─────────────────────────────────────────────
def load_flat(path):
    lib   = gdstk.read_gds(str(path))
    scale = lib.unit / 1e-6
    top   = lib.top_level()[0].copy("_flat", deep_copy=True)
    top.flatten()

    # 檢查坐標是否小了 10 倍：比較 bump 尺寸和坐標範圍
    # 例如：如果 bump W/H 在 10-100 範圍但坐標只有 1-100，可能坐標小了 10 倍
    all_pts = []
    for poly in top.polygons:
        if len(poly.points) >= 3:
            all_pts.extend(poly.points)

    import sys as _sys_scale
    return top, scale

    return top, scale

def get_chip_size(top, scale):
    """回傳含 SR+SL 的最大晶片邊框尺寸（取面積最大的 4-pt 矩形）"""
    best = None
    best_area = 0
    for poly in top.polygons:
        if len(poly.points) == 4:
            pts  = np.array(poly.points) * scale
            w    = float(pts[:,0].max() - pts[:,0].min())
            h    = float(pts[:,1].max() - pts[:,1].min())
            area = w * h
            if area > best_area:
                best_area = area
                best = (round(w, 4), round(h, 4))
    return best if best else (None, None)

def get_chip_bbox(top, scale):
    """回傳最大晶片矩形的 (xmin, ymin, xmax, ymax)"""
    best_bbox = None
    best_area = 0
    for poly in top.polygons:
        if len(poly.points) == 4:
            pts  = np.array(poly.points) * scale
            w    = float(pts[:,0].max() - pts[:,0].min())
            h    = float(pts[:,1].max() - pts[:,1].min())
            area = w * h
            if area > best_area:
                best_area = area
                best_bbox = (float(pts[:,0].min()), float(pts[:,1].min()),
                             float(pts[:,0].max()), float(pts[:,1].max()))
    return best_bbox

def get_amarks(top, scale):
    seen, result = set(), []
    for poly in top.polygons:
        if (poly.layer, poly.datatype) == AMARK_LAYER and len(poly.points) == 4:
            pts = np.array(poly.points) * scale
            w   = float(pts[:,0].max() - pts[:,0].min())
            h   = float(pts[:,1].max() - pts[:,1].min())
            if w * h < 1000:   # 過濾過小的多餘標記（< ~32×32 μm²）
                continue
            if max(w, h) > 2000:   # 排除 chip outline 等超大 polygon
                continue
            cx  = round(float(pts[:,0].mean()), 3)
            cy  = round(float(pts[:,1].mean()), 3)
            key = (cx, cy)
            if key not in seen:
                seen.add(key)
                result.append((cx, cy))
    result.sort()
    return result

def get_bumps(top, scale):
    """取出 bumps 並匹配 text 標注，按 NO 編號排序"""
    bumps = []
    for poly in top.polygons:
        if (poly.layer, poly.datatype) in BUMP_LAYERS:
            pts = np.array(poly.points) * scale
            cx  = round(float(pts[:,0].mean()), 4)
            cy  = round(float(pts[:,1].mean()), 4)
            w   = round(float(pts[:,0].max() - pts[:,0].min()), 4)
            h   = round(float(pts[:,1].max() - pts[:,1].min()), 4)
            if w > 5000:   # 排除 chip-size 等離群多邊形
                continue
            bumps.append({"cx": cx, "cy": cy, "w": w, "h": h, "number": "", "name": ""})

    # 收集 text labels
    num_labels, name_labels = [], []
    for lb in top.labels:
        lx = round(lb.origin[0] * scale, 4)
        ly = round(lb.origin[1] * scale, 4)
        if (lb.layer, lb.texttype) in TEXT_NUM_LAYERS and "NO" in lb.text.upper():
            num_labels.append((lx, ly, lb.text))
        if (lb.layer, lb.texttype) == TEXT_NAME_LAYER:
            name_labels.append((lx, ly, lb.text))

    # 最近鄰匹配
    if bumps:
        bxy = np.array([[b["cx"], b["cy"]] for b in bumps])
        for lx, ly, txt in num_labels:
            idx = int(np.argmin(np.linalg.norm(bxy - [lx, ly], axis=1)))
            bumps[idx]["number"] = txt
        for lx, ly, txt in name_labels:
            idx = int(np.argmin(np.linalg.norm(bxy - [lx, ly], axis=1)))
            bumps[idx]["name"] = txt

    # 按 NO 數字排序
    def _no(b):
        s = b["number"].upper().replace("NO", "").strip()
        return int(s) if s.isdigit() else 999999
    bumps.sort(key=_no)
    return bumps

def classify(b):
    """
    分類規則（模組預設佔位，run() 呼叫時會以空間密度邏輯動態覆寫）：
    - 密集單排（同 Y 行 > 10）+ 晶片下半部 → INPUT（靠近 IC 下半部 edge，1 排）
    - 密集單排（同 Y 行 > 10）+ 晶片上半部 → OUTPUT（靠近 IC 上半部 edge，1 排以上）
    - 其他（稀疏）→ DUMMY
    """
    return "DUMMY"

# ── Pad category colour palettes (shared across functions) ───────────────────
_PAL_INPUT  = ["4472C4", "70AD47", "ED7D31", "7030A0", "00B0F0"]
_PAL_OUTPUT = ["C00000", "FF7C80", "FF9900", "FFCC00", "FF66CC"]
_PAL_DUMMY  = ["7F7F7F", "A5A5A5", "BFBFBF"]

def _lighten(hex6: str, factor: float = 0.25) -> str:
    """Blend hex colour with white. factor=0→white, factor=1→original colour."""
    r=int(hex6[0:2],16); g=int(hex6[2:4],16); b=int(hex6[4:6],16)
    return (f"{int(r*factor+255*(1-factor)):02X}"
            f"{int(g*factor+255*(1-factor)):02X}"
            f"{int(b*factor+255*(1-factor)):02X}")

# ── 樣式 ─────────────────────────────────────────────────────────────────────
def _border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, r, c, val, bg="1F4E79", fg="FFFFFF", align="center"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(bold=True, color=fg, name="Calibri", size=11)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = _border("999999")
    return cell

def dat(ws, r, c, val, align="center", fmt=None, bold=False, bg=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _border()
    if fmt:
        cell.number_format = fmt
    # ✓ 統一所有單元格的字型（包含字體名稱，無論 bold 與否）
    cell.font = Font(bold=bold, size=11, name="Calibri")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell

# ── Sheet 1：PAD 清單（參考 SHIFT_7D5UM 格式）───────────────────────────────
def write_pad_sheet(ws, bumps, chip_w, chip_h, sheet_name):
    ws.title = sheet_name

    # 預設列高（影響所有未明確設定的資料列）
    ws.sheet_format.defaultRowHeight = 16
    ws.sheet_format.customHeight = True

    _info_fill = PatternFill("solid", fgColor="F2F2F2")

    # Row 1：Chip Size
    for c, v in [(1, "Chip_Size"), (2, "include SR+SL"),
                 (3, f"{chip_w} * {chip_h}" if chip_w else "—")]:
        cell = ws.cell(row=1, column=c, value=v)
        cell.font = Font(bold=(c == 1), size=12)
        cell.fill = _info_fill
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 20

    # Row 2：PAD Location
    cell2 = ws.cell(row=2, column=1, value="PAD Location")
    cell2.font = Font(bold=True, size=12)
    cell2.fill = _info_fill
    cell2.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 20

    # Row 3：欄位標題
    for c, h in enumerate(["No.", "Pad Name", "X-axis(um)", "Y-axis(um)",
                            "W(um)", "H(um)", "Area(um²)"], 1):
        hdr(ws, 3, c, h)
    ws.row_dimensions[3].height = 24

    # 面積統計（col I-K，rows 3-6）
    input_area  = sum(round(b["w"]) * round(b["h"]) for b in bumps if classify(b) == "INPUT")
    output_area = sum(round(b["w"]) * round(b["h"]) for b in bumps if classify(b) == "OUTPUT")
    dummy_area  = sum(round(b["w"]) * round(b["h"]) for b in bumps if classify(b) == "DUMMY")
    total_area  = input_area + output_area + dummy_area

    stats = [
        ("INPUT Bump area",  input_area,  "EBF3FB"),
        ("OUTPUT Bump area", output_area, "E2EFDA"),
        ("DUMMY Bump area",  dummy_area,  "FFF2CC"),
        ("Total Bump area",  total_area,  "D9E1F2"),
    ]
    for i, (label, val, bg) in enumerate(stats):
        row = 3 + i
        ws.row_dimensions[row].height = 20
        c_lbl = ws.cell(row=row, column=9, value=label)
        c_lbl.font      = Font(bold=True, size=11)
        c_lbl.fill      = PatternFill("solid", fgColor=bg)
        c_lbl.alignment = Alignment(horizontal="right", vertical="center")
        c_lbl.border    = _border("999999")
        c_val = ws.cell(row=row, column=10, value=val)
        c_val.font          = Font(size=11)
        c_val.number_format = "#,##0.00"
        c_val.alignment     = Alignment(horizontal="center", vertical="center")
        c_val.border        = _border("999999")
        c_val.fill          = PatternFill("solid", fgColor=bg)
        c_u = ws.cell(row=row, column=11, value="um²")
        c_u.font      = Font(size=11)
        c_u.alignment = Alignment(horizontal="left", vertical="center")

    # Row 4+：PAD 資料
    CAT_COLOR = {"INPUT": "EBF3FB", "OUTPUT": "E2EFDA", "DUMMY": "FFF2CC"}
    for i, b in enumerate(bumps, 4):
        cat = classify(b)
        bg  = CAT_COLOR.get(cat, "FFFFFF")
        dat(ws, i, 1, b["number"],                    bg=bg)
        dat(ws, i, 2, b["name"],       align="left",  bg=bg)
        dat(ws, i, 3, b["cx"],                        bg=bg)
        dat(ws, i, 4, b["cy"],                        bg=bg)
        dat(ws, i, 5, b["w"],                         bg=bg)
        dat(ws, i, 6, b["h"],                         bg=bg)
        dat(ws, i, 7, round(b["w"] * b["h"], 2),      bg=bg)

    # ── Bump Type Stat（col I-M，row 8 起）──
    from collections import Counter
    counts    = Counter((b["w"], b["h"]) for b in bumps)   # 用原始值，不四捨五入
    type_rows = sorted(counts.items(), key=lambda x: -x[1])
    stat_r    = 8
    for c, label in enumerate(["Bump Type", "W(um)", "H(um)", "Amount", "分類"], 1):
        hdr(ws, stat_r, 8 + c, label, bg="1F4E79")
    ws.row_dimensions[stat_r].height = 24

    # (W, H) → category lookup（每種尺寸取第一個 bump 的分類）
    _wh_to_cat: dict = {}
    for _b in bumps:
        _key = (_b["w"], _b["h"])
        if _key not in _wh_to_cat:
            _wh_to_cat[_key] = classify(_b)

    _CAT_BG  = {"INPUT": "BDD7EE", "OUTPUT": "C6EFCE", "DUMMY": "FFF2CC"}
    _CAT_FG  = {"INPUT": "1F4E79", "OUTPUT": "375623", "DUMMY": "7D6608"}

    for i, ((bw, bh), cnt) in enumerate(type_rows, stat_r + 1):
        _cat = _wh_to_cat.get((bw, bh), "DUMMY")
        _bg  = _CAT_BG[_cat]
        _fg  = _CAT_FG[_cat]
        ws.row_dimensions[i].height = 20
        dat(ws, i, 9,  i - stat_r, bold=True)
        dat(ws, i, 10, bw)
        dat(ws, i, 11, bh)
        dat(ws, i, 12, cnt)
        # 分類欄：顯示類別 + 色塊，使用者可直接在 Excel 修改
        cell = ws.cell(row=i, column=13, value=_cat)
        cell.font      = Font(bold=True, color=_fg, size=10,
                              name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=_bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        _s = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=_s, right=_s, top=_s, bottom=_s)

    # 欄寬
    for col, w in [("A",10),("B",26),("C",16),("D",16),("E",10),("F",10),("G",14),
                   ("H", 3),("I",22),("J",16),("K", 8),("L",16),("M",14)]:
        ws.column_dimensions[col].width = w

# ── 圖片生成 ─────────────────────────────────────────────────────────────────
def compute_bump_pitches(bumps, y_tol=2.0):
    """每個唯一 (W,H) 類型，按 y 行分組後計算最常見同排水平 pitch（μm）"""
    from collections import Counter
    grps = defaultdict(list)
    for b in bumps:
        grps[(round(b["w"]), round(b["h"]))].append(b)
    result = {}
    for key, grp in grps.items():
        y_groups = defaultdict(list)
        for b in grp:
            y_key = round(round(b["cy"] / y_tol) * y_tol, 2)
            y_groups[y_key].append(b["cx"])
        all_diffs = []
        for xs in y_groups.values():
            xs_s = sorted(xs)
            for i in range(len(xs_s) - 1):
                d = round(xs_s[i+1] - xs_s[i], 1)
                if 0 < d < 500:
                    all_diffs.append(d)
        if all_diffs:
            result[key] = Counter(all_diffs).most_common(1)[0][0]
    return result


def compute_vertical_pitch(cat_bumps, h_pitch=None, min_diff=10.0):
    """計算 bump 群的垂直行間距：先按 x 欄分組，再取欄內 y 差值的眾數。
    可避免對角排列的小 y-步距（如 1.6um）干擾結果。"""
    from collections import Counter, defaultdict as _ddd
    if len(cat_bumps) < 2:
        return None
    # 若未提供 H-pitch，先快速估算
    if h_pitch is None:
        _yg = _ddd(list)
        for b in cat_bumps:
            _yg[round(round(b["cy"]/2)*2, 2)].append(b["cx"])
        _xd = []
        for xs in _yg.values():
            xs_s = sorted(xs)
            for i in range(len(xs_s)-1):
                d = round(xs_s[i+1]-xs_s[i], 1)
                if 5 < d < 500:
                    _xd.append(d)
        h_pitch = Counter(round(d, 1) for d in _xd).most_common(1)[0][0] if _xd else 44.0
    # 按 x 欄分組（以 h_pitch/2 為 tolerance）
    x_tol = h_pitch / 2
    x_grps = _ddd(list)
    for b in cat_bumps:
        xk = round(round(b["cx"] / x_tol) * x_tol, 1)
        x_grps[xk].append(b["cy"])
    # 收集欄內 y 差值
    all_ydiffs = []
    for ys_list in x_grps.values():
        ys = sorted(ys_list)
        for i in range(len(ys)-1):
            d = round(ys[i+1]-ys[i], 1)
            if d >= min_diff:
                all_ydiffs.append(d)
    if not all_ydiffs:
        return None
    return Counter(all_ydiffs).most_common(1)[0][0]


def compute_all_pitches(cat_bumps, tol=0.5, y_tol=2.0):
    """一個 bump 類別的所有不重複水平間距（先按 y 行分組，再計算同行 x 差值，sorted）"""
    if len(cat_bumps) < 2:
        return []
    # 按 y 分組
    y_groups = defaultdict(list)
    for b in cat_bumps:
        y_key = round(round(b["cy"] / y_tol) * y_tol, 2)
        y_groups[y_key].append(round(b["cx"], 2))
    all_diffs = []
    for xs in y_groups.values():
        xs_s = sorted(set(xs))
        for i in range(len(xs_s) - 1):
            d = round(xs_s[i+1] - xs_s[i], 2)
            if d > 0:
                all_diffs.append(d)
    if not all_diffs:
        return []
    unique_sorted = sorted(set(round(d, 2) for d in all_diffs))
    merged = []
    for v in unique_sorted:
        if not merged or v - merged[-1] > tol:
            merged.append(round(v, 2))
    return merged


def _draw_amark_symbol(ax, cx, cy, size, color="red", lw=0.8, aspect=1.0):
    """AMARK 標準符號：外框正方形 + 4 個直角三角形（四象限各一）
    aspect: x-stretch factor to compensate embedding distortion (embed_sy/embed_sx)"""
    Ox, Oy = size/2 * aspect, size/2
    Cx, Cy = size*0.125 * aspect, size*0.125
    Tx, Ty = size*0.375 * aspect, size*0.375
    ax.add_patch(plt.Polygon(
        [[cx-Ox,cy-Oy],[cx+Ox,cy-Oy],[cx+Ox,cy+Oy],[cx-Ox,cy+Oy]],
        closed=True, facecolor="none", edgecolor=color, linewidth=lw, zorder=10))
    for _sx, _sy in [(1,1),(-1,1),(-1,-1),(1,-1)]:
        tri = [[cx+_sx*Cx, cy+_sy*Cy],
               [cx+_sx*Tx, cy+_sy*Cy],
               [cx+_sx*Cx, cy+_sy*Ty]]
        ax.add_patch(plt.Polygon(tri, closed=True, facecolor="none",
                                  edgecolor=color, linewidth=lw*0.8, zorder=10))


def _collect_amark_polys(top, scale):
    """Return list of pts arrays for all AMARK polygons (outer rect + inner cross).
    Outer: AMARK_LAYER 4-pt rectangles.
    Inner: TEXT_NAME_LAYER polygons whose centroid falls inside an outer bbox.
    排除 chip outline 等超大 polygon（適用 AMARK_LAYER 與 chip_size 同 layer 的情況，如 JD9365T）。"""
    outer_list, inner_candidates = [], []
    for poly in top.polygons:
        pts = np.array(poly.points) * scale
        key = (poly.layer, poly.datatype)
        bw = float(pts[:,0].max() - pts[:,0].min())
        bh = float(pts[:,1].max() - pts[:,1].min())
        if max(bw, bh) > 2000:   # 排除 chip outline 等超大 polygon
            continue
        if key == AMARK_LAYER and len(poly.points) == 4:
            outer_list.append(pts)
        elif key == TEXT_NAME_LAYER:
            inner_candidates.append(pts)
    result = list(outer_list)
    for ipts in inner_candidates:
        cx = float(ipts[:,0].mean()); cy = float(ipts[:,1].mean())
        for opts in outer_list:
            if (float(opts[:,0].min()) <= cx <= float(opts[:,0].max()) and
                    float(opts[:,1].min()) <= cy <= float(opts[:,1].max())):
                result.append(ipts)
                break
    return result


def render_chip_overview(top, scale, chip_w, chip_h, bumps=None, bump_symbols=None) -> bytes:
    """生成晶片全覽圖（白底，bump 黃色，外框黑線，寬高標注，pitch 標注）回傳 PNG bytes"""
    _trapz = getattr(np, "trapezoid", None) or getattr(np, "trapz")

    all_pts = np.vstack([np.array(p.points) * scale for p in top.polygons
                         if len(p.points) >= 3])
    xmin = float(all_pts[:,0].min()); xmax = float(all_pts[:,0].max())
    ymin = float(all_pts[:,1].min()); ymax = float(all_pts[:,1].max())
    W = xmax - xmin; H = ymax - ymin

    fig_w = 50
    fig_h = 2.0
    fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=150)
    ax.set_facecolor("white"); fig.patch.set_facecolor("white")

    # 晶片輪廓（polygon，面積 > 1e7）
    for poly in top.polygons:
        pts = np.array(poly.points) * scale
        area = abs(float(_trapz(pts[:,1], pts[:,0])))
        if area > 1e7:
            ax.add_patch(plt.Polygon(pts, closed=True, facecolor="none",
                                     edgecolor="black", linewidth=0.5, zorder=1))
    for apts in _collect_amark_polys(top, scale):
        ax.add_patch(plt.Polygon(apts, closed=True, facecolor="none",
                                 edgecolor="black", linewidth=0.5, zorder=3))

    # 各類型 pad 著色（from bumps list）
    if bumps:
        _ov_in  = sorted(set((round(b["w"]),round(b["h"])) for b in bumps if classify(b)=="INPUT"))
        _ov_out = sorted(set((round(b["w"]),round(b["h"])) for b in bumps if classify(b)=="OUTPUT"))
        _ov_dum = sorted(set((round(b["w"]),round(b["h"])) for b in bumps if classify(b)=="DUMMY"))
        _cat_color = {}
        for i,k in enumerate(_ov_in):  _cat_color[("INPUT",  k)] = "#"+_PAL_INPUT [i%len(_PAL_INPUT )]
        for i,k in enumerate(_ov_out): _cat_color[("OUTPUT", k)] = "#"+_PAL_OUTPUT[i%len(_PAL_OUTPUT)]
        for i,k in enumerate(_ov_dum): _cat_color[("DUMMY",  k)] = "#"+_PAL_DUMMY [i%len(_PAL_DUMMY )]
        for b in bumps:
            cat = classify(b)
            c = _cat_color.get((cat, (round(b["w"]), round(b["h"]))))
            if c:
                ax.add_patch(plt.Rectangle(
                    (b["cx"] - b["w"]/2, b["cy"] - b["h"]/2), b["w"], b["h"],
                    facecolor=c, edgecolor="none", zorder=2, alpha=0.8))

    pad = W * 0.02
    ax.set_xlim(xmin - pad, xmax + pad)
    ax.set_ylim(ymin - H * 0.05, ymax + H * 0.18)
    ax.set_aspect("equal"); ax.axis("off")

    # 寬高標注
    arr_y = ymax + H * 0.06
    ax.annotate("", xy=(xmax, arr_y), xytext=(xmin, arr_y),
                arrowprops=dict(arrowstyle="<->", color="red", lw=0.6))
    ax.text((xmin+xmax)/2, arr_y + H*0.03,
            f"{chip_w}um (SR + SL)", ha="center", va="bottom",
            fontsize=9, color="red", fontweight="bold")
    arr_x = xmin - W * 0.015
    ax.annotate("", xy=(arr_x, ymax), xytext=(arr_x, ymin),
                arrowprops=dict(arrowstyle="<->", color="red", lw=0.6))
    ax.text(arr_x - W*0.005, (ymin+ymax)/2,
            f"{chip_h}um\n(SR + SL)", ha="right", va="center",
            fontsize=9, color="red", fontweight="bold", rotation=90)

    # IC 中心準星（箭頭樣式）
    cx = (xmin+xmax)/2; cy = (ymin+ymax)/2
    ax.annotate("", xy=(cx + W*0.025, cy), xytext=(cx - W*0.025, cy),
                arrowprops=dict(arrowstyle="<->", color="red", lw=0.6, mutation_scale=6), zorder=10)
    ax.annotate("", xy=(cx, cy + H*0.18), xytext=(cx, cy - H*0.18),
                arrowprops=dict(arrowstyle="<->", color="red", lw=0.6, mutation_scale=6), zorder=10)
    ax.text(cx + W*0.027, cy, "X", fontsize=10, color="red", ha="left", va="center", fontweight="bold")
    ax.text(cx, cy + H*0.20, "Y", fontsize=10, color="red", ha="center", va="bottom", fontweight="bold")
    ax.text(cx + W*0.006, cy - H*0.04, "(0,0)", fontsize=8, color="red", ha="left", va="top")

    # ── Option A: Symbol labels on bump groups ────────────────────────────────
    if bumps and bump_symbols:
        _mid_y = (ymin + ymax) / 2
        _sym_rep = {}  # symbol → (cx, cy, h)
        for b in bumps:
            _cat = classify(b)
            _key = (_cat, (round(b["w"]), round(b["h"])))
            _sym = bump_symbols.get(_key)
            if _sym and (_sym not in _sym_rep or b["cx"] < _sym_rep[_sym][0]):
                _sym_rep[_sym] = (b["cx"], b["cy"], b["h"])
        for _sym, (_lx, _lcy, _lh) in sorted(_sym_rep.items()):
            _is_bot = _lcy < _mid_y
            _ty = (_lcy + _lh/2 + H*0.01) if _is_bot else (_lcy - _lh/2 - H*0.01)
            ax.text(_lx, _ty, _sym, ha="center",
                    va="bottom" if _is_bot else "top",
                    fontsize=6.5, fontweight="bold", color="black", zorder=15,
                    bbox=dict(boxstyle="square,pad=0.08", fc="white", ec="none", alpha=0.85))
    # AMARK 輪廓保留，D1/D2 文字標籤移至放大圖顯示

    # ── Option B: Edge distance arrows (A4/A5/A7) ────────────────────────────
    if bumps:
        _in_list  = [b for b in bumps if classify(b) == "INPUT"]
        _out_list = [b for b in bumps if classify(b) == "OUTPUT"]
        _arr_kw = dict(arrowstyle="<->", color="#0070C0", lw=0.6)
        _tk_kw  = dict(fontsize=6, color="#0070C0", fontweight="bold")

        if _in_list:
            # A4: BUMP to IC bottom edge (vertical arrow at leftmost INPUT bump)
            _left_b   = min(_in_list, key=lambda b: b["cx"])
            _a4_x     = _left_b["cx"]
            _a4_b_top = _left_b["cy"] - _left_b["h"]/2
            ax.annotate("", xy=(_a4_x, _a4_b_top), xytext=(_a4_x, ymin),
                        arrowprops=_arr_kw, zorder=12)
            ax.text(_a4_x + W*0.005, (ymin + _a4_b_top)/2, "A4",
                    ha="left", va="center", **_tk_kw)

            # A5: BUMP to IC side edge (horizontal arrow below chip)
            _a5_lx = min(b["cx"] - b["w"]/2 for b in _in_list)
            _a5_y  = ymin - H * 0.028
            ax.annotate("", xy=(_a5_lx, _a5_y), xytext=(xmin, _a5_y),
                        arrowprops=_arr_kw, zorder=12)
            ax.text((xmin + _a5_lx)/2, _a5_y - H*0.008, "A5",
                    ha="center", va="top", **_tk_kw)

        # A7: OUTPUT-to-INPUT gap (vertical arrow at IC center)
        if _in_list and _out_list:
            _in_top  = max(b["cy"] + b["h"]/2 for b in _in_list)
            _cxm     = (xmin + xmax) / 2
            _near_o  = [b for b in _out_list if abs(b["cx"] - _cxm) <= (xmax-xmin)/4]
            _out_bot = min(b["cy"] - b["h"]/2 for b in (_near_o or _out_list))
            _a7_x    = _cxm  # IC center
            ax.annotate("", xy=(_a7_x, _out_bot), xytext=(_a7_x, _in_top),
                        arrowprops=_arr_kw, zorder=12)
            ax.text(_a7_x + W*0.008, (_in_top + _out_bot)/2, "A7",
                    ha="left", va="center", **_tk_kw)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def render_chip_sections(top, scale, bumps=None, n=7) -> list:
    """將晶片水平等分為 n 份，各自渲染截面 PNG，回傳 list of bytes（左→右）"""
    _trapz = getattr(np, "trapezoid", None) or getattr(np, "trapz")

    all_pts = np.vstack([np.array(p.points) * scale for p in top.polygons
                         if len(p.points) >= 3])
    xmin = float(all_pts[:,0].min()); xmax = float(all_pts[:,0].max())
    ymin = float(all_pts[:,1].min()); ymax = float(all_pts[:,1].max())
    W = xmax - xmin; H = ymax - ymin

    # 預先收集多邊形
    chip_polys, bump_polys = [], []
    for poly in top.polygons:
        pts = np.array(poly.points) * scale
        key = (poly.layer, poly.datatype)
        pw  = float(pts[:,0].max() - pts[:,0].min())
        area = abs(float(_trapz(pts[:,1], pts[:,0])))
        if area > 1e7:
            chip_polys.append(pts)
        elif key in BUMP_LAYERS and pw < 5000:
            bump_polys.append(pts)
    amark_polys = _collect_amark_polys(top, scale)

    sec_w = W / n
    pad_y = H * 0.12
    results = []

    for i in range(n):
        x0 = xmin + i * sec_w
        x1 = xmin + (i + 1) * sec_w
        mx = sec_w * 0.04          # 小橫向留白

        aspect = sec_w / (H + 2 * pad_y)
        fig_w = 8
        fig_h = max(1.5, fig_w / aspect)
        fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=150)
        ax.set_facecolor("white"); fig.patch.set_facecolor("white")

        for pts in chip_polys:
            ax.add_patch(plt.Polygon(pts, closed=True, facecolor="none",
                                     edgecolor="black", linewidth=1.2, zorder=1))
        for pts in bump_polys:
            ax.add_patch(plt.Polygon(pts, closed=True, facecolor="white",
                                     edgecolor="black", linewidth=0.3, zorder=2))
        for apts in amark_polys:
            ax.add_patch(plt.Polygon(apts, closed=True, facecolor="none",
                                     edgecolor="black", linewidth=1.0, zorder=3))

        ax.set_xlim(x0 - mx, x1 + mx)
        ax.set_ylim(ymin - pad_y, ymax + pad_y)
        ax.set_aspect("equal"); ax.axis("off")

        # 截面編號（中心=4，兩旁各3）
        label = f"Section {i+1} / {n}"
        ax.text((x0+x1)/2, ymax + pad_y * 0.55, label,
                ha="center", va="bottom", fontsize=9,
                color="#1F4E79", fontweight="bold")

        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
        plt.close(fig)
        buf.seek(0)
        results.append(buf.read())

    return results


def compute_auto_window_um(bumps, chip_w):
    """Compute recommended window_um based on dominant bump pitch and chip width."""
    from collections import Counter as _C
    if not bumps or not chip_w:
        return max(int(round(chip_w * 0.10 / 500)) * 500 if chip_w else 3000, 1500)
    wh_cnt = _C((round(b["w"]), round(b["h"])) for b in bumps)
    rep_wh = wh_cnt.most_common(1)[0][0]
    rep = sorted([b for b in bumps if round(b["w"]) == rep_wh[0] and round(b["h"]) == rep_wh[1]],
                 key=lambda b: b["cx"])
    target = chip_w * 0.10
    if len(rep) >= 2:
        dxs = [rep[i+1]["cx"] - rep[i]["cx"] for i in range(min(50, len(rep)-1))]
        dxs_pos = [d for d in dxs if 0 < d < 500]
        if dxs_pos:
            med = sorted(dxs_pos)[len(dxs_pos) // 2]
            target = med * 90
    lo = max(chip_w * 0.09, 2500)
    hi = chip_w * 0.25
    return int(round(max(lo, min(hi, target)) / 500)) * 500


def render_key_sections(top, scale, bumps=None, window_um=3500, bump_symbols=None) -> bytes:
    """5 個重點截面拼接：左EDGE + A1/A2/A3 樣本 + 右EDGE → 單張 PNG bytes

    Args:
        window_um: 各區域寬度（µm），預設 3500，A2 Center 固定 ±1750 不受影響
    """
    _trapz = getattr(np, "trapezoid", None) or getattr(np, "trapz")

    # 確保使用全局 classify（反映 overrides）
    global classify

    all_pts = np.vstack([np.array(p.points) * scale for p in top.polygons if len(p.points) >= 3])
    xmin, xmax = float(all_pts[:,0].min()), float(all_pts[:,0].max())
    ymin, ymax = float(all_pts[:,1].min()), float(all_pts[:,1].max())
    H = ymax - ymin

    # 各區域寬度（user 可定制，預設 3500μm）；A2 Center 固定 ±1750
    # regions 格式：(label, x0, x1)
    _sec_w = float(window_um)
    chip_bbox_pre = get_chip_bbox(top, scale)
    if chip_bbox_pre:
        _bx0r, _by0r, _bx1r, _by1r = chip_bbox_pre
    else:
        _bx0r, _bx1r = xmin, xmax

    # EDGE sections
    _s0_x0, _s0_x1 = _bx0r, _bx0r + _sec_w
    _s4_x0, _s4_x1 = _bx1r - _sec_w, _bx1r

    # INPUT bump types present in each EDGE section
    _in_ks = [b for b in (bumps or []) if classify(b) == "INPUT"]
    _types_l = {(round(b["w"]),round(b["h"])) for b in _in_ks if _s0_x0 <= b["cx"] <= _s0_x1}
    _types_r = {(round(b["w"]),round(b["h"])) for b in _in_ks if _s4_x0 <= b["cx"] <= _s4_x1}

    # Left transition: first bump (left→right) beyond section 0 whose type is NOT in left edge
    _trans_l = next(
        (b["cx"] for b in sorted(_in_ks, key=lambda b: b["cx"])
         if b["cx"] > _s0_x1 and (round(b["w"]),round(b["h"])) not in _types_l),
        None)
    # Right transition: first bump (right→left) before section 4 whose type is NOT in right edge
    _trans_r = next(
        (b["cx"] for b in sorted(_in_ks, key=lambda b: -b["cx"])
         if b["cx"] < _s4_x0 and (round(b["w"]),round(b["h"])) not in _types_r),
        None)

    # AI fallback: find leftmost adjacent INPUT pair beyond section 0 (for A4 h-gap)
    _best_in_cx = None
    _in_ks_cx = sorted([b for b in _in_ks if b["cx"] > _s0_x1], key=lambda b: b["cx"])
    for _ii in range(len(_in_ks_cx) - 1):
        _bi, _bj = _in_ks_cx[_ii], _in_ks_cx[_ii+1]
        if abs(_bi["cy"] - _bj["cy"]) < 5:
            _gi = _bj["cx"] - _bj["w"]/2 - (_bi["cx"] + _bi["w"]/2)
            if 0 < _gi < 200:
                _best_in_cx = (_bi["cx"] + _bj["cx"]) / 2
                break

    # AI fallback: find rightmost adjacent OUTPUT pair before section 4 (for B7 h-gap)
    _out_ks_all = [b for b in (bumps or []) if classify(b) == "OUTPUT"]
    _best_out_cx = None
    _out_ks_cx = sorted([b for b in _out_ks_all if b["cx"] < _s4_x0], key=lambda b: -b["cx"])
    for _ii in range(len(_out_ks_cx) - 1):
        _bi, _bj = _out_ks_cx[_ii], _out_ks_cx[_ii+1]  # right→left
        if abs(_bi["cy"] - _bj["cy"]) < 5:
            _go = _bi["cx"] - _bi["w"]/2 - (_bj["cx"] + _bj["w"]/2)
            if 0 < _go < 300:
                _best_out_cx = (_bi["cx"] + _bj["cx"]) / 2
                break

    # Section 1: transition (type change) or AI center on adjacent INPUT pair
    if _trans_l is not None:
        _s1_x0 = max(_s0_x1 + 50, _trans_l - _sec_w * 0.3)
        _s1_x1 = _s1_x0 + _sec_w
    elif _best_in_cx is not None:
        _s1_x0 = _best_in_cx - _sec_w / 2
        _s1_x1 = _best_in_cx + _sec_w / 2
    else:
        _s1_x0 = _s0_x1; _s1_x1 = _s1_x0 + _sec_w

    # Section 3: transition or AI center on adjacent OUTPUT pair
    if _trans_r is not None:
        _s3_x1 = min(_s4_x0 - 50, _trans_r + _sec_w * 0.3)
        _s3_x0 = _s3_x1 - _sec_w
    elif _best_out_cx is not None:
        _s3_x0 = _best_out_cx - _sec_w / 2
        _s3_x1 = _best_out_cx + _sec_w / 2
    else:
        _s3_x1 = _s4_x0; _s3_x0 = _s3_x1 - _sec_w

    # Section 2: A2 Center — 固定 ±1750μm，剩餘中間均歸此區
    _mid = (_bx0r + _bx1r) / 2
    _ctr_x0, _ctr_x1 = _mid - 1750, _mid + 1750

    def _fmt(v): return f"{v:.0f}µm"
    regions = [
        (f"#1  {_fmt(_s0_x0)}~{_fmt(_s0_x1)}",                    _s0_x0, _s0_x1),
        (f"#2  {_fmt(_s1_x0)}~{_fmt(_s1_x1)}",                    _s1_x0, _s1_x1),
        (f"#3  A2 Center\n{_fmt(_ctr_x0)}~{_fmt(_ctr_x1)}",         _ctr_x0, _ctr_x1),
        (f"#4  {_fmt(_s3_x0)}~{_fmt(_s3_x1)}",                    _s3_x0, _s3_x1),
        (f"#5  {_fmt(_s4_x0)}~{_fmt(_s4_x1)}",                    _s4_x0, _s4_x1),
    ]

    # 預先收集多邊形
    chip_polys, bump_polys = [], []
    for poly in top.polygons:
        pts = np.array(poly.points) * scale
        k   = (poly.layer, poly.datatype)
        pw  = float(pts[:,0].max() - pts[:,0].min())
        area = abs(float(_trapz(pts[:,1], pts[:,0])))
        if area > 1e7:
            chip_polys.append(pts)
        elif k in BUMP_LAYERS and pw < 5000:
            bump_polys.append(pts)
    amark_polys = _collect_amark_polys(top, scale)

    # ── Pad type colour maps ──────────────────────────────────────────────────
    _in_keys_ks  = sorted(set((round(b["w"]),round(b["h"])) for b in (bumps or []) if classify(b)=="INPUT"))
    _out_keys_ks = sorted(set((round(b["w"]),round(b["h"])) for b in (bumps or []) if classify(b)=="OUTPUT"))
    _dum_keys_ks = sorted(set((round(b["w"]),round(b["h"])) for b in (bumps or []) if classify(b)=="DUMMY"))
    _in_color_ks  = {k:"#"+_PAL_INPUT [i%len(_PAL_INPUT )] for i,k in enumerate(_in_keys_ks)}
    _out_color_ks = {k:"#"+_PAL_OUTPUT[i%len(_PAL_OUTPUT)] for i,k in enumerate(_out_keys_ks)}
    _dum_color_ks = {k:"#"+_PAL_DUMMY [i%len(_PAL_DUMMY )] for i,k in enumerate(_dum_keys_ks)}

    # ── Per-category bumps & edge distances ──────────────────────────────────
    chip_bbox_ks = get_chip_bbox(top, scale)
    in_bumps_ks  = [b for b in (bumps or []) if classify(b)=="INPUT"]
    out_bumps_ks = [b for b in (bumps or []) if classify(b)=="OUTPUT"]
    dum_bumps_ks = [b for b in (bumps or []) if classify(b)=="DUMMY"]
    n_in  = len(_in_keys_ks)
    n_out = len(_out_keys_ks)
    n_dum = len(_dum_keys_ks)

    # Pre-compute OUTPUT h-pitch by grouping bumps into rows (same cy ±5um)
    _out_hp_ks = None
    if out_bumps_ks:
        from collections import defaultdict as _dd_out
        _out_row_g = _dd_out(list)
        for _b in out_bumps_ks:
            _out_row_g[round(_b["cy"] / 5) * 5].append(_b)
        _best_out_row = max(_out_row_g.values(), key=len)
        if len(_best_out_row) >= 2:
            _br_s = sorted(_best_out_row, key=lambda b: b["cx"])
            _odxs = [_br_s[j+1]["cx"] - _br_s[j]["cx"] for j in range(len(_br_s)-1)]
            _odxs_pos = [d for d in _odxs if 0 < d < 300]
            if _odxs_pos:
                _out_hp_ks = sorted(_odxs_pos)[len(_odxs_pos) // 2]
        if _out_hp_ks is None:
            _oa = sorted(out_bumps_ks, key=lambda b: b["cx"])
            _fb = [_oa[j+1]["cx"]-_oa[j]["cx"] for j in range(min(30, len(_oa)-1))]
            _fb_pos = [d for d in _fb if 0 < d < 300]
            if _fb_pos:
                _out_hp_ks = sorted(_fb_pos)[len(_fb_pos)//2]

    # INPUT A4/A5
    _a4_bottom_y = min(b["cy"]-b["h"]/2 for b in in_bumps_ks) if in_bumps_ks else None
    _a5_left_x   = min(b["cx"]-b["w"]/2 for b in in_bumps_ks) if in_bumps_ks else None
    # A5 畫在 i==0（Left EDGE），5000μm 視窗已涵蓋 IC 左緣和第一個 INPUT pad

    # A2：OUTPUT 下緣 - INPUT 上緣間距的中點（只計算 OUTPUT 和 INPUT，忽略 DUMMY）
    # 計算：OUTPUT 最下方 cy - h/2，INPUT 最上方 cy + h/2，取中點
    _a2_sec = None; _a2_info = None   # (bump_cx, output_bottom_y, input_top_y, mid_y)
    if in_bumps_ks and out_bumps_ks and chip_bbox_ks:
        _input_top_y = max(b["cy"] + b["h"]/2 for b in in_bumps_ks)

        # Chip 中心點 X 座標（與 B4 CALC 一致）
        _chip_center_x_a2 = (chip_bbox_ks[0] + chip_bbox_ks[2]) / 2
        _chip_width_a2 = chip_bbox_ks[2] - chip_bbox_ks[0]
        _center_tolerance_a2 = _chip_width_a2 / 4

        # OUTPUT 最下排：只用靠近 chip 中心的 OUTPUT bumps
        _near_out_bumps_a2 = [b for b in out_bumps_ks if abs(b["cx"] - _chip_center_x_a2) <= _center_tolerance_a2]
        if _near_out_bumps_a2:
            _output_bottom_y = min(b["cy"] - b["h"]/2 for b in _near_out_bumps_a2)
        else:
            _output_bottom_y = min(b["cy"] - b["h"]/2 for b in out_bumps_ks)

        # 計算中點（無論坐標順序）
        _a2_mid_y = (_input_top_y + _output_bottom_y) / 2
        # 取 INPUT bumps 中心 X 座標平均值
        _a2_cx = sum(b["cx"] for b in in_bumps_ks) / len(in_bumps_ks)
        _bx0, _, _bx1, _by1 = chip_bbox_ks
        _a2_sec = None
        for _ri, (_rl, _rx0, _rx1) in enumerate(regions):
            if _rx0 <= _a2_cx <= _rx1:
                _a2_sec = _ri
                break
        if _a2_sec is None:
            _a2_sec = len(regions) - 1  # fallback
        _a2_info = (_a2_cx, _output_bottom_y, _input_top_y, _a2_mid_y)

    # C3（垂直，類似 B2）：DUMMY bump 頂部 → IC 上緣
    # C4（水平，類似 B3）：DUMMY bump 最靠近 IC 左/右緣，強制歸到 Left/Right EDGE 截面
    _c3_sec = None; _c3_info = None   # (bump_cx, pad_top_y, ic_top_y)
    _c4_sec = None; _c4_info = None   # (pad_edge_x, ic_edge_x, bump_cy)
    if dum_bumps_ks and chip_bbox_ks:
        _bx0,_by0,_bx1,_by1 = chip_bbox_ks
        _dum_inside = [b for b in dum_bumps_ks
                       if _bx0 < b["cx"] < _bx1 and _by0 < b["cy"] < _by1]
        if _dum_inside:
            def _sec_of(cx):
                for _ri, (_rl, _rx0, _rx1) in enumerate(regions):
                    if _rx0 <= cx <= _rx1:
                        return _ri
                return len(regions) - 1

            # C3：取左上方第一組 DUMMY bump 為標示起始點
            # 左上象限（cx < chip 中心 X，cy > chip 中心 Y）→ 取最左（次取最上）
            def _top_gap(b):
                g = _by1 - (b["cy"] + b["h"]/2)
                return g if g >= 0 else float("inf")
            _chip_cx_ks = (_bx0 + _bx1) / 2
            _chip_cy_ks = (_by0 + _by1) / 2
            _top_left_pool = [b for b in _dum_inside
                              if b["cx"] < _chip_cx_ks and b["cy"] > _chip_cy_ks]
            if _top_left_pool:
                _c3b = min(_top_left_pool, key=lambda b: (b["cx"], -b["cy"]))  # 最左，次取最上
            else:
                # fallback：整體最左的 DUMMY bump
                _c3b = min(_dum_inside, key=lambda b: b["cx"])
            _c3_sec = _sec_of(_c3b["cx"])
            if _top_gap(_c3b) < float("inf"):
                _c3_info = (_c3b["cx"], _c3b["cy"]+_c3b["h"]/2, _by1)

            # C4：最靠近 IC 左/右緣，強制歸 Left(0) 或 Right(4) EDGE 截面
            def _horiz_gap(b):
                gl = b["cx"]-b["w"]/2-_bx0
                gr = _bx1-(b["cx"]+b["w"]/2)
                pos = [g for g in [gl, gr] if g >= 0]
                return min(pos) if pos else float("inf")
            _c4b = min(_dum_inside, key=_horiz_gap)
            if _horiz_gap(_c4b) < float("inf"):
                _gl = _c4b["cx"]-_c4b["w"]/2-_bx0
                _gr = _bx1-(_c4b["cx"]+_c4b["w"]/2)
                if _gl >= 0 and (_gr < 0 or _gl <= _gr):
                    _c4_info = (_c4b["cx"]-_c4b["w"]/2, _bx0, _c4b["cy"])  # left→IC left
                    _c4_sec  = 0   # Left EDGE section
                else:
                    _c4_info = (_c4b["cx"]+_c4b["w"]/2, _bx1, _c4b["cy"])  # right→IC right
                    _c4_sec  = len(regions) - 1  # Right EDGE section

    n = len(regions)
    pad_y = H * 0.20          # 下方留更多空間給 annotation

    # B3 位置：Right EDGE 截面中 AMARK 頂部 y（箭頭放在 AMARK 上方）
    _amark_rgt_top = None
    _rx0 = regions[4][1]   # Right EDGE section 左邊界
    for _apts in amark_polys:
        if float(_apts[:,0].mean()) > _rx0:
            _t = float(_apts[:,1].max())
            if _amark_rgt_top is None or _t > _amark_rgt_top:
                _amark_rgt_top = _t

    # 預先群聚 AMARK polygons → gap-based：相鄰 poly cx 差 > 500 um 開新 cluster
    _am_polys_sorted = sorted(amark_polys, key=lambda p: float(p[:,0].mean()))
    _am_marks = []  # list of [mark_cx, mark_cy, [polys]]
    _AM_GAP = 500.0  # um；同一 AMARK mark 內 poly 間距不超過此值
    for _apts in _am_polys_sorted:
        _acx = float(_apts[:,0].mean())
        _acy = float(_apts[:,1].mean())
        if _am_marks and (_acx - float(_am_marks[-1][2][-1][:,0].mean())) < _AM_GAP:
            _am_marks[-1][2].append(_apts)
            _all_pts = np.vstack(_am_marks[-1][2])
            _am_marks[-1][0] = float(_all_pts[:,0].mean())
            _am_marks[-1][1] = float(_all_pts[:,1].mean())
        else:
            _am_marks.append([_acx, _acy, [_apts]])

    # 以 cluster bounding box 過濾雜散小聚合（< 40um）及 chip outline 等超大聚合（> 2000um）
    _am_marks_valid = []
    for _mk in _am_marks:
        _all_mk = np.vstack(_mk[2])
        _mk_bw = float(_all_mk[:,0].max() - _all_mk[:,0].min())
        _mk_bh = float(_all_mk[:,1].max() - _all_mk[:,1].min())
        if 40 <= max(_mk_bw, _mk_bh) <= 2000:
            _am_marks_valid.append(_mk)
    _am_marks = _am_marks_valid

    # 嵌入尺寸 4000×430 → 9.30:1；PNG 同比例避免變形
    _DPI   = 300
    fig_w  = 4000 / _DPI      # 13.33"
    fig_h  = 430  / _DPI      # 1.43"

    fig = plt.figure(figsize=(fig_w, fig_h), dpi=_DPI)
    fig.patch.set_facecolor("white")
    # left/right 留白給 B3/C4 等文字延伸到 chip edge 外
    gs = fig.add_gridspec(1, n, wspace=0.08,  # 增加間距以容納省略號
                           left=0.03, right=0.91, top=0.82, bottom=0.02)

    _in_hgap_drawn = False   # h-gap (A{n_in+3}) 只畫在第一個有 INPUT 對的 section
    _in_wspan_drawn = False  # W-span (A{n_in+5}) 只畫在第二個有 INPUT 對的 section

    for i, (label, x0, x1) in enumerate(regions):
        ax = fig.add_subplot(gs[0, i])
        ax.set_facecolor("white")

        for pts in chip_polys:
            ax.add_patch(plt.Polygon(pts, closed=True, facecolor="none",
                                      edgecolor="black", linewidth=0.4, zorder=1))
        for pts in bump_polys:
            ax.add_patch(plt.Polygon(pts, closed=True, facecolor="white",
                                      edgecolor="black", linewidth=0.2, zorder=2))
        # AMARK 多邊形 + D1/D2 標籤（每個 mark 群聚後只標一次）
        for _ami, (_mk_cx, _mk_cy, _mk_polys) in enumerate(_am_marks):
            if x0 - H*0.5 < _mk_cx < x1 + H*0.5:
                for apts in _mk_polys:
                    ax.add_patch(plt.Polygon(apts, closed=True, facecolor="none",
                                             edgecolor="black", linewidth=0.4, zorder=3))
                ax.text(_mk_cx, _mk_cy, f"D{_ami+1}", ha="center", va="center",
                        fontsize=6, fontweight="bold", color="red", zorder=15)

        # ── pad 著色 + Symbol 標籤 (two-pass: draw pads, then stagger labels) ──
        _mid_y_ks = (ymin + ymax) / 2
        _sym_reps  = {}  # sym → (leftmost_cx, cy, h, is_bot)

        # Pass 1: draw pads, collect leftmost representative per symbol
        for cat, color_map in [("INPUT",_in_color_ks),("OUTPUT",_out_color_ks),("DUMMY",_dum_color_ks)]:
            for b in (bumps or []):
                if classify(b) == cat and x0 <= b["cx"] <= x1:
                    _bkey = (round(b["w"]), round(b["h"]))
                    c = color_map.get(_bkey)
                    if c:
                        ax.add_patch(plt.Rectangle(
                            (b["cx"]-b["w"]/2, b["cy"]-b["h"]/2), b["w"], b["h"],
                            facecolor=c, edgecolor="none", zorder=3, alpha=0.85))
                    if bump_symbols:
                        _sym = bump_symbols.get((cat, _bkey))
                        if _sym and (_sym not in _sym_reps or b["cx"] < _sym_reps[_sym][0]):
                            _is_bot = b["cy"] < _mid_y_ks
                            _sym_reps[_sym] = (b["cx"], b["cy"], b["h"], _is_bot)

        # Pass 2: collision-based shift (only push right when labels would overlap)
        if bump_symbols and _sym_reps:
            _bot_grp = sorted([(s,d) for s,d in _sym_reps.items() if     d[3]], key=lambda x: x[1][0])
            _top_grp = sorted([(s,d) for s,d in _sym_reps.items() if not d[3]], key=lambda x: x[1][0])
            _lbl_w = H * 0.12  # estimated label width in um
            for _grp, _is_bot in [(_bot_grp, True), (_top_grp, False)]:
                _placed = []  # placed x positions
                for _sym, (_bcx, _bcy, _bh, _) in _grp:
                    _base = _bcy + _bh/2 if _is_bot else _bcy - _bh/2
                    _lty  = _base + H*0.025
                    _ltx  = _bcx
                    for _px in _placed:
                        if abs(_ltx - _px) < _lbl_w:
                            _ltx = _px + _lbl_w
                    _placed.append(_ltx)
                    ax.text(_ltx, _lty, _sym, ha="center",
                            va="bottom" if _is_bot else "top",
                            fontsize=5.5, fontweight="bold", color="black",
                            zorder=16,
                            bbox=dict(boxstyle="square,pad=0.06", fc="white",
                                      ec="none", alpha=0.85))

        _cjk = {"fontfamily": "Microsoft JhengHei"}
        _sw = x1 - x0

        # text_y: 覆寫文字 y 位置（None = 使用箭頭中點）
        _txt_bbox = dict(boxstyle="square,pad=0.05", fc="white", ec="none", alpha=0.92)
        def _arrow_v(ax_obj, x_ann, y_pad, y_ic, sym, col, text_y=None):
            if abs(y_pad - y_ic) < 0.01: return
            ax_obj.annotate("", xy=(x_ann, y_ic), xytext=(x_ann, y_pad),
                           arrowprops=dict(arrowstyle="<->", color=col, lw=0.5, mutation_scale=6),
                           annotation_clip=False)
            _ty = text_y if text_y is not None else (y_ic+y_pad)/2
            ax_obj.text(x_ann + _sw*0.04, _ty, sym,
                       ha="left", va="center", fontsize=6.5, color=col,
                       fontweight="bold", zorder=15, clip_on=False,
                       bbox=_txt_bbox, **_cjk)

        # text_outside=True：文字放在 x_to 外側（chip edge 外）
        def _arrow_h(ax_obj, x_from, x_to, y_ann, sym, col, text_outside=False):
            if abs(x_from - x_to) < 0.1: return
            ax_obj.annotate("", xy=(x_to, y_ann), xytext=(x_from, y_ann),
                           arrowprops=dict(arrowstyle="<->", color=col, lw=0.5, mutation_scale=6),
                           annotation_clip=False)
            if text_outside:
                _tx = x_to + _sw*0.06 if x_to > x_from else x_to - _sw*0.06
                _ha = "left" if x_to > x_from else "right"
                ax_obj.text(_tx, y_ann, sym,
                           ha=_ha, va="center", fontsize=6.5, color=col,
                           fontweight="bold", zorder=15, clip_on=False,
                           bbox=_txt_bbox, **_cjk)
            else:
                ax_obj.text((x_from+x_to)/2, y_ann - pad_y*0.2, sym,
                           ha="center", va="top", fontsize=6.5, color=col,
                           fontweight="bold", zorder=15, clip_on=False,
                           bbox=_txt_bbox, **_cjk)

        if chip_bbox_ks:
            cbxmin, cbymin, cbxmax, cbymax = chip_bbox_ks
            col_a = "#1F3864"   # 深藍，比 INPUT pad fill (#4472C4) 深，確保對比
            col_b = "#7B0000"   # 深棗紅，比 OUTPUT pad fill (#C00000) 深，確保對比

            # A4：i==0（Left EDGE），INPUT 垂直，pad 下方 → IC 下緣；x 對齊最左 INPUT bump
            if i == 0 and _a4_bottom_y is not None:
                _in_sec0 = [b for b in in_bumps_ks if x0 <= b["cx"] <= x1]
                _a4_x = min(_in_sec0, key=lambda b: b["cx"])["cx"] if _in_sec0 else \
                        min(in_bumps_ks, key=lambda b: b["cx"])["cx"]
                _arrow_v(ax, _a4_x, _a4_bottom_y, cbymin, f"A{n_in+1}", col_a)

            # A5：i==0，section 0 xlim 已擴大涵蓋兩端點，IC 左緣 → 第一個 INPUT pad 左緣
            if i == 0 and _a5_left_x is not None and _a5_left_x > cbxmin:
                _arrow_h(ax, cbxmin, _a5_left_x, cbymin - pad_y*0.55, f"A{n_in+2}", col_a)

            # A2：OUTPUT 下緣 - INPUT 上緣間距中點
            if _a2_info and _a2_sec == i:
                col_a2 = "#" + _PAL_INPUT[0]  # 使用 INPUT 顏色（靠近 INPUT）
                _a2_cx, _a2_bot_y, _a2_top_y, _a2_mid_y = _a2_info
                import sys as _sys_dbg
                print(f"[A2 INFO] _a2_cx={_a2_cx}, _a2_bot_y={_a2_bot_y}, _a2_top_y={_a2_top_y}, dist={abs(_a2_top_y-_a2_bot_y)}, section={i}/{len(regions)}", file=_sys_dbg.stderr)

                # A7：OI gap，symbol 與 INFORMATION 表格一致
                _arrow_v(ax, _a2_cx, _a2_bot_y, _a2_top_y, f"A{n_in+4}", col_a2,
                         text_y=_a2_mid_y)

            # B{2n+1}：i==2（A2 Center），OUTPUT 垂直，OUTPUT pad 頂部 → IC 上緣
            if i == 2 and out_bumps_ks:
                top_out = max(b["cy"]+b["h"]/2 for b in out_bumps_ks)
                _arrow_v(ax, x0 + _sw*0.15, top_out, cbymax, f"B{2*n_out+1}", col_b)

            # B{2n+2}：i==4（Right EDGE），OUTPUT 水平 → IC 右緣，文字延伸到 chip edge 外
            if i == len(regions)-1 and out_bumps_ks:
                rgt_b = max(out_bumps_ks, key=lambda b: b["cx"]+b["w"]/2)
                _b3_y = (_amark_rgt_top + pad_y*0.40) if _amark_rgt_top is not None \
                        else cbymax + pad_y*0.25
                _arrow_h(ax, rgt_b["cx"]+rgt_b["w"]/2, cbxmax,
                         _b3_y, f"B{2*n_out+2}", col_b, text_outside=True)

            # C{2n+1}（垂直）：DUMMY pad 頂部 → IC 上緣
            if _c3_info and _c3_sec == i and n_dum > 0:
                col_c = "#2F2F2F"   # 深灰，比 DUMMY pad fill (#7F7F7F) 深
                _c3_cx, _c3_pad_top, _c3_ic_top = _c3_info
                _arrow_v(ax, _c3_cx + _sw*0.10, _c3_pad_top, _c3_ic_top,
                         f"C{2*n_dum+1}", col_c, text_y=(_c3_pad_top + _c3_ic_top) / 2)

            # C{2n+2}（水平）：DUMMY pad 側緣 → IC 側緣
            if _c4_info and _c4_sec == i and n_dum > 0:
                col_c = "#2F2F2F"
                _c4_pad_x, _c4_ic_x, _c4_bcy = _c4_info
                _arrow_h(ax, _c4_pad_x, _c4_ic_x, _c4_bcy,
                         f"C{2*n_dum+2}", col_c, text_outside=True)

            # A6：水平 bump 間距（INPUT 相鄰兩個 bump 之間的 gap）
            # 在此截面中找兩個相鄰 INPUT bump，畫出右邊緣→下一個左邊緣的箭頭
            _in_sec = [b for b in in_bumps_ks if x0 <= b["cx"] <= x1]
            if _in_sec and in_bumps_ks and i != 0:  # skip section 0 to avoid overlap with A4
                # 計算全局 h_pitch（INPUT）
                _inp_hp_ks = None
                if h_pitches_ks_in := {k: v for k, v in
                                        {(round(b["w"]),round(b["h"])): None
                                         for b in in_bumps_ks}.items()}:
                    from collections import Counter as _C6
                    _wh_cnt = _C6((round(b["w"]),round(b["h"])) for b in in_bumps_ks)
                    _rep_k  = _wh_cnt.most_common(1)[0][0]
                    # use pre-computed h_pitches from outer scope via closest match
                    _in_ks_sorted_cx = sorted(
                        [b for b in in_bumps_ks if (round(b["w"]),round(b["h"])) == _rep_k],
                        key=lambda b: b["cx"])
                    if len(_in_ks_sorted_cx) >= 2:
                        _dxs = [_in_ks_sorted_cx[j+1]["cx"] - _in_ks_sorted_cx[j]["cx"]
                                for j in range(min(20, len(_in_ks_sorted_cx)-1))]
                        _dxs_pos = [d for d in _dxs if 0 < d < 200]
                        if _dxs_pos:
                            _inp_hp_ks = sorted(_dxs_pos)[len(_dxs_pos)//2]  # median

                if _inp_hp_ks:
                    # 找此截面中同一排（相同 cy）相鄰兩個 INPUT bump
                    _in_sec_s = sorted(_in_sec, key=lambda b: b["cx"])
                    _a6_pair = None
                    for _jj in range(len(_in_sec_s)-1):
                        _b1, _b2 = _in_sec_s[_jj], _in_sec_s[_jj+1]
                        if abs(_b1["cy"] - _b2["cy"]) < 5:
                            _gap6 = _b2["cx"] - _b2["w"]/2 - (_b1["cx"] + _b1["w"]/2)
                            if 0 < _gap6 < _inp_hp_ks:
                                _a6_pair = (_b1, _b2, _gap6)
                                break
                    if _a6_pair:
                        _b1, _b2, _gap6 = _a6_pair
                        if not _in_hgap_drawn:
                            # h-gap：畫在第一個有 INPUT 對的 section
                            _in_hgap_drawn = True
                            _a6_y = _b1["cy"] - _b1["h"]/2 - pad_y*0.28
                            _x_from = _b1["cx"] + _b1["w"]/2
                            _x_to   = _b2["cx"] - _b2["w"]/2
                            _arrow_h(ax, _x_from, _x_to, _a6_y, f"A{n_in+3}", col_a)
                        elif not _in_wspan_drawn:
                            # W-span：畫在第二個有 INPUT 對的 section（與 h-gap 不同 subplot）
                            _in_wspan_drawn = True
                            _aw_y = _b1["cy"] - _b1["h"]/2 - pad_y*0.28
                            _arrow_h(ax, _b1["cx"]-_b1["w"]/2, _b1["cx"]+_b1["w"]/2,
                                     _aw_y, f"A{n_in+5}", col_a)

        # B{2n+3}：OUTPUT h-gap（水平，section i==2）— row-based grouping for staggered grids
        if chip_bbox_ks and n_out > 0 and i == 2:
            _out_sec_i = [b for b in out_bumps_ks if x0 <= b["cx"] <= x1]
            if _out_sec_i:
                from collections import defaultdict as _dd_outrow
                _out_row_map2 = _dd_outrow(list)
                for _ob in _out_sec_i:
                    _out_row_map2[round(_ob["cy"] / 5) * 5].append(_ob)
                _b7_pair = None
                _hp_thresh = _out_hp_ks if _out_hp_ks else 200
                # 從最上排開始找（避免在 OI gap 區域與 A{n_in+4} 文字重疊）
                for _rcy_k in sorted(_out_row_map2.keys(), reverse=True):
                    _row_bumps = _out_row_map2[_rcy_k]
                    if len(_row_bumps) >= 2:
                        _row_s = sorted(_row_bumps, key=lambda b: b["cx"])
                        for _jj in range(len(_row_s) - 1):
                            _ob1, _ob2 = _row_s[_jj], _row_s[_jj+1]
                            _gb7 = _ob2["cx"] - _ob2["w"]/2 - (_ob1["cx"] + _ob1["w"]/2)
                            if 0 < _gb7 < _hp_thresh:
                                _b7_pair = (_ob1, _ob2)
                                break
                    if _b7_pair:
                        break
                if _b7_pair:
                    _ob1, _ob2 = _b7_pair
                    _b7_y = _ob1["cy"] + _ob1["h"]/2 + pad_y*0.12   # 上排 TOP 上方，遠離 OI gap
                    _arrow_h(ax, _ob1["cx"]+_ob1["w"]/2, _ob2["cx"]-_ob2["w"]/2,
                             _b7_y, f"B{2*n_out+3}", col_b)

        # B{2n+4}：OUTPUT v-pitch gap（垂直，section i==2）— 最上方相鄰兩排，固定 x 右側
        if chip_bbox_ks and n_out > 0 and i == 2:
            _out_v_sec = [b for b in out_bumps_ks if x0 <= b["cx"] <= x1]
            if _out_v_sec:
                # 按 cy 降序找最上方相鄰兩排的 gap
                _out_cy_desc = sorted(_out_v_sec, key=lambda b: -b["cy"])
                _b6_pair = None
                for _jj in range(len(_out_cy_desc) - 1):
                    _vb_hi, _vb_lo = _out_cy_desc[_jj], _out_cy_desc[_jj+1]
                    _gv = _vb_hi["cy"] - _vb_hi["h"]/2 - (_vb_lo["cy"] + _vb_lo["h"]/2)
                    if 0 < _gv < 200:
                        _b6_pair = (_vb_lo, _vb_hi)   # lo=bottom, hi=top
                        break
                if _b6_pair:
                    _vb_lo, _vb_hi = _b6_pair
                    _b6_x = x0 + _sw * 0.82   # 固定在 section 右側，遠離 B{2n+1} 左側
                    _arrow_v(ax, _b6_x,
                             _vb_lo["cy"] + _vb_lo["h"]/2,
                             _vb_hi["cy"] - _vb_hi["h"]/2,
                             f"B{2*n_out+4}", col_b)

        ax.set_xlim(x0, x1)
        ax.set_ylim(ymin - pad_y, ymax + pad_y * 0.3)
        ax.axis("off")
        ax.set_title(label, fontsize=6, color="#1F4E79", fontweight="bold", pad=2,
                     fontfamily="Microsoft JhengHei")

        # 右側虛線分隔（最後一張除外）
        if i < n - 1:
            ax.spines["right"].set_visible(True)
            ax.spines["right"].set_color("#AAAAAA")
            ax.spines["right"].set_linewidth(1.0)
            ax.spines["right"].set_linestyle("dashed")

    # ── 在截面邊界處添加省略號 ────────────────────────────────────────────────
    # 計算每個截面的相對位置（gridspec 5 列均勻分布）
    # left=0.03, right=0.91, 總寬度=0.88
    _left_margin = 0.03
    _right_margin = 1 - 0.91
    _usable_width = 0.88
    _col_width = _usable_width / n
    for _i in range(n - 1):
        # 第 i 列和第 i+1 列之間的中點
        _x_pos = _left_margin + (_i + 1) * _col_width
        fig.text(_x_pos, 0.50, "···", ha="center", va="center",
                fontsize=8, color="#999999", fontweight="normal")

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=300, facecolor="white")  # key sections
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def render_amark_detail(top, scale, ax_um, ay_um, label) -> bytes:
    """Amark detail image with engineering dimension annotations"""
    # Collect nearby polygons (relative to amark center)
    raw = []
    for poly in top.polygons:
        pts = np.array(poly.points) * scale
        cx  = float(pts[:,0].mean()); cy = float(pts[:,1].mean())
        pw  = float(pts[:,0].max()-pts[:,0].min())
        if abs(cx-ax_um) < 200 and abs(cy-ay_um) < 200 and pw < 200:
            rel = pts - np.array([ax_um, ay_um])
            raw.append((poly.layer, poly.datatype, len(poly.points), rel))

    # Deduplicate (flatten may double-reference sub-cells)
    seen, unique = set(), []
    for layer, dtype, n, pts in raw:
        key = (layer, dtype, n, tuple(map(tuple, np.round(pts, 1).tolist())))
        if key not in seen:
            seen.add(key); unique.append((layer, dtype, n, pts))

    outer = None; inner = []
    _amark_polys = [(n, pts) for layer, dtype, n, pts in unique if (layer, dtype) == AMARK_LAYER]
    _other_polys = [(n, pts) for layer, dtype, n, pts in unique
                    if (layer, dtype) != AMARK_LAYER and (layer, dtype) not in BUMP_LAYERS]
    if _amark_polys:
        # 最大面積 polygon = 外框；其餘 = 十字 inner
        _amark_polys.sort(
            key=lambda x: (x[1][:,0].max()-x[1][:,0].min()) * (x[1][:,1].max()-x[1][:,1].min()),
            reverse=True)
        outer = _amark_polys[0][1]
        inner = _amark_polys[1:] + _other_polys
    else:
        inner = _other_polys

    if outer is None:
        return b""

    # 從實際外框計算 O（半邊長）
    O = float(max(outer[:,0].max()-outer[:,0].min(),
                  outer[:,1].max()-outer[:,1].min())) / 2
    MARGIN = round(O * 0.4)

    # 是否為十字（兩個方向不同的矩形 inner）
    _rects = [(n, pts) for n, pts in inner if n == 4]
    _has_h = any((pts[:,0].max()-pts[:,0].min()) > (pts[:,1].max()-pts[:,1].min())
                 for _, pts in _rects)
    _has_v = any((pts[:,1].max()-pts[:,1].min()) > (pts[:,0].max()-pts[:,0].min())
                 for _, pts in _rects)
    is_cross = ((_has_h and _has_v) or any(n >= 10 for n, _ in inner))

    fig, ap = plt.subplots(figsize=(5.5, 5.5), dpi=120)
    ap.set_facecolor("white"); fig.patch.set_facecolor("white")
    ap.set_aspect("equal"); ap.axis("off")
    ap.set_xlim(-O - MARGIN, O + MARGIN)
    ap.set_ylim(-O - MARGIN - 4, O + MARGIN + 6)

    # Draw outer box
    ap.add_patch(plt.Polygon(outer, closed=True, facecolor="none",
                             edgecolor="black", linewidth=1.5, zorder=2))
    # Draw inner shapes
    for n, pts in inner:
        ap.add_patch(plt.Polygon(pts, closed=True, facecolor="none",
                                edgecolor="black", linewidth=1.0, zorder=3))

    # Red label at top-left
    ap.text(-O - MARGIN + 1, O + MARGIN + 4, label,
            ha='left', va='top', fontsize=12, fontweight='bold', color='red')

    # ── Engineering dimension helpers ─────────────────────────────────────────
    AP = dict(arrowstyle='<->', color='black', lw=0.8, mutation_scale=9)

    def dim_h(x1, x2, y_base, y_line, text, fs=8.5):
        """Horizontal dim: double arrow from x1→x2 at y=y_line, ext lines from y_base"""
        if abs(y_line - y_base) > 0.3:
            ap.plot([x1, x1], [y_base, y_line], 'k-', lw=0.5, zorder=5)
            ap.plot([x2, x2], [y_base, y_line], 'k-', lw=0.5, zorder=5)
        ap.annotate('', xy=(x2, y_line), xytext=(x1, y_line), arrowprops=AP)
        sign = 1 if y_line >= y_base else -1
        ap.text((x1+x2)/2, y_line + sign*2.5, text,
               ha='center', va='bottom' if sign > 0 else 'top', fontsize=fs)

    def dim_v(y1, y2, x_base, x_line, text, fs=8.5):
        """Vertical dim: double arrow from y1→y2 at x=x_line, ext lines from x_base"""
        if abs(x_line - x_base) > 0.3:
            ap.plot([x_base, x_line], [y1, y1], 'k-', lw=0.5, zorder=5)
            ap.plot([x_base, x_line], [y2, y2], 'k-', lw=0.5, zorder=5)
        ap.annotate('', xy=(x_line, y2), xytext=(x_line, y1), arrowprops=AP)
        sign = 1 if x_line >= x_base else -1
        ap.text(x_line + sign*2.5, (y1+y2)/2, text,
               ha='left' if sign > 0 else 'right', va='center',
               fontsize=fs, rotation=90)

    if is_cross:
        # 從 inner 矩形自動推算 cross 尺寸
        _h_rect = next((pts for n, pts in _rects
                        if (pts[:,0].max()-pts[:,0].min()) > (pts[:,1].max()-pts[:,1].min())), None)
        _v_rect = next((pts for n, pts in _rects
                        if (pts[:,1].max()-pts[:,1].min()) > (pts[:,0].max()-pts[:,0].min())), None)
        if _h_rect is not None and _v_rect is not None:
            _arm_w  = float(min(_h_rect[:,1].max()-_h_rect[:,1].min(),
                                _v_rect[:,0].max()-_v_rect[:,0].min()))  # 臂寬
            _arm_len= float(max(_h_rect[:,0].max()-_h_rect[:,0].min(),
                                _v_rect[:,1].max()-_v_rect[:,1].min()))  # 臂長
        else:
            _arm_w, _arm_len = O * 0.5, O * 1.5
        A  = round(_arm_w  / 2, 2)   # 臂寬半值
        B  = round(_arm_len/ 2, 2)   # 臂頂距中心
        _gap = round(O - B, 2)        # 外框到臂頂的間距
        _ow  = round(O * 2, 1)        # 外框尺寸

        dim_h(-O, O, O,  O+MARGIN*0.6, str(_ow))
        dim_v(-O, O, -O, -O-MARGIN*0.6, str(_ow))

        dim_v(B, O, A, A + MARGIN*0.4, str(_gap), fs=7.5)
        dim_h(-O, -B, -A, -A - MARGIN*0.4, str(_gap), fs=7.5)

        dim_h(-A, A, B, B + MARGIN*0.3, str(round(_arm_w,1)), fs=7.5)
        dim_v(-A, A, B, B + MARGIN*0.3, str(round(_arm_w,1)), fs=7.5)

        _arm_half = round(B - A, 2)
        dim_v(A, B, A, A + MARGIN*0.3, str(_arm_half), fs=7.5)
        dim_h(A, B, A, A + MARGIN*0.3, str(_arm_half), fs=7.5)

    else:
        # Triangles: 依實際 outer 尺寸等比計算（_ow=100 時對應 C=12.5, T=37.5, 邊=37.5, 中=25）
        _ow   = round(O * 2, 1)
        C     = round(_ow * 0.125, 3)
        T     = round(_ow * 0.375, 3)
        _edge = round(O - C, 2)
        _mid  = round(2 * C, 2)

        # Outside: 外框尺寸
        dim_h(-O, O, O,  O+15, str(_ow))
        dim_v(-O, O, -O, -O-15, str(_ow))

        # Bottom horizontal chain: _edge | _mid | _edge
        BOT = -O - 11
        dim_h(-O, -C, -O, BOT, str(_edge), fs=7.5)
        dim_h(-C,  C, -O, BOT, str(_mid),  fs=7.5)
        dim_h(C,   O, -O, BOT, str(_edge), fs=7.5)

        # Right vertical chain: _edge | _mid | _edge
        RGT = O + 11
        dim_v(-O, -C, O, RGT, str(_edge), fs=7.5)
        dim_v(-C,  C, O, RGT, str(_mid),  fs=7.5)
        dim_v(C,   O, O, RGT, str(_edge), fs=7.5)

        # Triangle leg annotation at top-right
        dim_h(C, T, C, C+9, str(_mid), fs=7.5)
        dim_v(C, T, C, C+9, str(_mid), fs=7.5)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def render_top_view_bumps(top, scale, bumps=None) -> bytes:
    """左EDGE 3500μm 截面 face-up 頂視圖（無標注，等比例，含標題）"""
    _trapz = getattr(np, "trapezoid", None) or getattr(np, "trapz")
    all_pts = np.vstack([np.array(p.points) * scale for p in top.polygons if len(p.points) >= 3])
    xmin = float(all_pts[:,0].min()); xmax = float(all_pts[:,0].max())
    ymin = float(all_pts[:,1].min()); ymax = float(all_pts[:,1].max())
    H = ymax - ymin

    view_x0 = xmin - 50
    view_x1 = xmin + 3500 + 50
    pad_y = H * 0.08

    _DPI = 200
    fig_w = (view_x1 - view_x0) / 1000 * 3.5
    fig_h = (H + pad_y * 2) / 1000 * 3.5
    fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=_DPI)
    ax.set_facecolor("white"); fig.patch.set_facecolor("white")

    # chip outline
    for poly in top.polygons:
        pts = np.array(poly.points) * scale
        area = abs(float(_trapz(pts[:,1], pts[:,0])))
        if area > 1e7:
            ax.add_patch(plt.Polygon(pts, closed=True, facecolor="none",
                                     edgecolor="black", linewidth=0.4, zorder=1))
    for apts in _collect_amark_polys(top, scale):
        acx = float(apts[:,0].mean())
        if view_x0 <= acx <= view_x1:
            ax.add_patch(plt.Polygon(apts, closed=True, facecolor="none",
                                     edgecolor="black", linewidth=0.4, zorder=3))

    # pad coloring
    if bumps:
        _ov_in  = sorted(set((round(b["w"]),round(b["h"])) for b in bumps if classify(b)=="INPUT"))
        _ov_out = sorted(set((round(b["w"]),round(b["h"])) for b in bumps if classify(b)=="OUTPUT"))
        _ov_dum = sorted(set((round(b["w"]),round(b["h"])) for b in bumps if classify(b)=="DUMMY"))
        _cat_color = {}
        for i,k in enumerate(_ov_in):  _cat_color[("INPUT",  k)] = "#"+_PAL_INPUT [i%len(_PAL_INPUT )]
        for i,k in enumerate(_ov_out): _cat_color[("OUTPUT", k)] = "#"+_PAL_OUTPUT[i%len(_PAL_OUTPUT)]
        for i,k in enumerate(_ov_dum): _cat_color[("DUMMY",  k)] = "#"+_PAL_DUMMY [i%len(_PAL_DUMMY )]
        for b in bumps:
            if not (view_x0 <= b["cx"] <= view_x1):
                continue
            cat = classify(b)
            c = _cat_color.get((cat, (round(b["w"]), round(b["h"]))))
            if c:
                ax.add_patch(plt.Rectangle(
                    (b["cx"] - b["w"]/2, b["cy"] - b["h"]/2), b["w"], b["h"],
                    facecolor=c, edgecolor="none", zorder=2, alpha=0.85))

    ax.set_xlim(view_x0, view_x1)
    ax.set_ylim(ymin - pad_y, ymax + pad_y)
    ax.set_aspect("equal"); ax.axis("off")
    ax.set_title("Top view: Bump face-up", fontsize=8, fontweight="bold",
                 color="#1F4E79", pad=3)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=_DPI, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _embed_image(ws, png_bytes, anchor_cell, width_px=None, height_px=None):
    """將 PNG bytes 嵌入工作表"""
    if not png_bytes:
        return
    img = XLImage(io.BytesIO(png_bytes))
    if width_px:
        img.width  = width_px
    if height_px:
        img.height = height_px
    img.anchor = anchor_cell
    ws.add_image(img)


# ── Sheet 2：INFORMATION ──────────────────────────────────────────────────────
def write_information(ws, bumps, chip_w, chip_h, amarks, top=None, scale=None, window_um=3500):
    # 統計各類 bump 的尺寸組合
    input_types  = defaultdict(lambda: {"count": 0, "nos": []})
    output_types = defaultdict(lambda: {"count": 0, "nos": []})
    dummy_types  = defaultdict(lambda: {"count": 0, "nos": []})

    for b in bumps:
        cat  = classify(b)
        key  = (round(b["w"]), round(b["h"]))
        no   = b["number"]
        if cat == "INPUT":
            input_types[key]["count"] += 1
            input_types[key]["nos"].append(no)
        elif cat == "OUTPUT":
            output_types[key]["count"] += 1
            output_types[key]["nos"].append(no)
        else:
            dummy_types[key]["count"] += 1
            dummy_types[key]["nos"].append(no)

    def no_range(nos):
        if not nos: return ""
        return f"{nos[0]}~{nos[-1]}"

    # ── 晶片全覽圖（嵌入頂部）──
    OVW_ROWS  = 10   # 全覽圖列數
    KEY_ROWS  = 11   # 重點截面拼接圖列數
    TOP_ROWS  = 5    # Top view 圖列數（縮小後僅需 5 行）

    if top is not None and scale is not None:
        _pre_in_s  = sorted(input_types.keys())
        _pre_out_s = sorted(output_types.keys())
        _pre_dum_s = sorted(dummy_types.keys())
        _bsyms = {}
        for _i, _k in enumerate(_pre_in_s):  _bsyms[("INPUT",  _k)] = f"A{_i+1}"
        for _i, _k in enumerate(_pre_out_s): _bsyms[("OUTPUT", _k)] = f"B{2*_i+1}"
        for _i, _k in enumerate(_pre_dum_s): _bsyms[("DUMMY",  _k)] = f"C{2*_i+1}"
        png = render_chip_overview(top, scale, chip_w, chip_h, bumps=bumps, bump_symbols=_bsyms)
        _embed_image(ws, png, "A1", width_px=2800, height_px=140)
    for ri in range(1, OVW_ROWS + 1):
        ws.row_dimensions[ri].height = 11

    # ── 5 重點截面拼接圖（左EDGE + A1/A2/A3 + 右EDGE）──
    if top is not None and scale is not None:
        key_png = render_key_sections(top, scale, bumps=bumps, window_um=window_um, bump_symbols=_bsyms)
        key_r = OVW_ROWS + 1
        for ri in range(key_r, key_r + KEY_ROWS):
            ws.row_dimensions[ri].height = 21
        _embed_image(ws, key_png, f"A{key_r}", width_px=2800, height_px=300)

        # ── Top view: Bump face-up（A22 起始，縮圖）──
        top_r = key_r + KEY_ROWS  # row 22
        _top_view_path = Path(__file__).parent / "input" / "top_view_face.png"
        top_png = _top_view_path.read_bytes() if _top_view_path.exists() else None
        for ri in range(top_r, top_r + TOP_ROWS):
            ws.row_dimensions[ri].height = 26
        _embed_image(ws, top_png, f"A{top_r}", width_px=220, height_px=172)
        CHIP_IMG_ROWS = OVW_ROWS + KEY_ROWS + TOP_ROWS
    else:
        CHIP_IMG_ROWS = OVW_ROWS

    # ── Condition + Notes（圖片下方）──
    r = CHIP_IMG_ROWS + 1
    _blk  = lambda t, bold=False: TextBlock(InlineFont(b=bold, sz=11, color="000000"), t)
    _red  = lambda t, bold=False: TextBlock(InlineFont(b=bold, sz=11, color="FF0000"), t)
    c1 = ws.cell(row=r, column=5)
    c1.value = CellRichText(_blk("Condition-1 "), _red("(+/- 2um)", bold=True), _blk(" temp. compensation  Unit=um"))
    c1.alignment = Alignment(vertical="center")

    note_rows = [
        CellRichText(_blk("Note1: Chip size includes "), _red("60um"), _blk(" scribe line Before Wafer sawing")),
        CellRichText(_blk("Note2: BUMP to IC edge "), _red("include scribe line")),
        CellRichText(_blk("Note3: Chip size One side Tolerance "), _red("30um"), _blk(" After Wafer sawing")),
    ]
    for j, note_rt in enumerate(note_rows):
        c = ws.cell(row=r + j, column=9)
        c.value = note_rt
        c.alignment = Alignment(vertical="center")

    r += len(note_rows) + 1

    # ── PAD 分類 headers ──────────────────────────────────────────────────────
    # 4 cols per section: Symbol | Size | Number | Tolerance
    # AMARK: 3 cols: Symbol | Size | Tolerance
    ISEC, OSEC, DSEC, ASEC = 1, 5, 9, 13

    for col, label, bg in [(ISEC,"INPUT PAD","1F4E79"), (OSEC,"OUTPUT PAD","375623"),
                           (DSEC,"DUMMY PAD","7F6000"), (ASEC,"AMARK","833C11")]:
        c = ws.cell(row=r, column=col, value=label)
        c.font = Font(bold=True, color="FFFFFF", size=13)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")

    r += 1
    for col, label in [
        (ISEC,  "Symbol"),(ISEC+1, "Size"),(ISEC+2,"Number"),(ISEC+3,"Tolerance"),
        (OSEC,  "Symbol"),(OSEC+1, "Size"),(OSEC+2,"Number"),(OSEC+3,"Tolerance"),
        (DSEC,  "Symbol"),(DSEC+1, "Size"),(DSEC+2,"Number"),(DSEC+3,"Tolerance"),
        (ASEC,  "Symbol"),(ASEC+1, "Size"),(ASEC+2,"Tolerance"),
    ]:
        hdr(ws, r, col, label, bg="2E75B6")

    r += 1

    def _no_range_str(nos):
        nums = [int(n.upper().replace("NO","").strip())
                for n in nos if n.upper().replace("NO","").strip().isdigit()]
        if not nums: return ""
        return f"NO.{min(nums)}~NO.{max(nums)}"

    # ── 先分類再預計算每個 (W,H) 的 H-pitch、V-pitch ────────────────────────────────
    # 按 classify 結果分組，為每個分類單獨計算 pitch
    in_bumps_list = [b for b in bumps if classify(b) == "INPUT"]
    out_bumps_list = [b for b in bumps if classify(b) == "OUTPUT"]
    dum_bumps_list = [b for b in bumps if classify(b) == "DUMMY"]

    import sys as _sys_pitch
    print(f"\n[PITCH CALC] in={len(in_bumps_list)}, out={len(out_bumps_list)}, dum={len(dum_bumps_list)}", file=_sys_pitch.stderr)

    # 為 INPUT、OUTPUT、DUMMY 分別計算 pitch
    h_pitches_in = compute_bump_pitches(in_bumps_list)
    h_pitches_out = compute_bump_pitches(out_bumps_list)
    h_pitches_dum = compute_bump_pitches(dum_bumps_list)

    print(f"[PITCH CALC] h_pitches_in={h_pitches_in}", file=_sys_pitch.stderr)
    print(f"[PITCH CALC] h_pitches_out={h_pitches_out}", file=_sys_pitch.stderr)
    print(f"[PITCH CALC] h_pitches_dum={h_pitches_dum}", file=_sys_pitch.stderr)

    # 合併三個字典（OUTPUT 優先級最高，其次 INPUT，DUMMY 最低）
    h_pitches = {**h_pitches_dum, **h_pitches_in, **h_pitches_out}

    # 計算 V-pitch 時也按分類計算
    key_bumps_in  = defaultdict(list)
    key_bumps_out = defaultdict(list)
    key_bumps_dum = defaultdict(list)
    for b in in_bumps_list:
        key_bumps_in[(round(b["w"]), round(b["h"]))].append(b)
    for b in out_bumps_list:
        key_bumps_out[(round(b["w"]), round(b["h"]))].append(b)
    for b in dum_bumps_list:
        key_bumps_dum[(round(b["w"]), round(b["h"]))].append(b)

    v_pitches_in = {k: compute_vertical_pitch(v, h_pitch=h_pitches.get(k)) for k, v in key_bumps_in.items()}
    v_pitches_out = {k: compute_vertical_pitch(v, h_pitch=h_pitches.get(k)) for k, v in key_bumps_out.items()}
    v_pitches_dum = {k: compute_vertical_pitch(v, h_pitch=h_pitches.get(k)) for k, v in key_bumps_dum.items()}

    print(f"[PITCH CALC] v_pitches_in={v_pitches_in}", file=_sys_pitch.stderr)
    print(f"[PITCH CALC] v_pitches_out={v_pitches_out}", file=_sys_pitch.stderr)
    print(f"[PITCH CALC] v_pitches_dum={v_pitches_dum}", file=_sys_pitch.stderr)

    # 合併 V-pitch 字典
    v_pitches = {**v_pitches_dum, **v_pitches_in, **v_pitches_out}

    def _fmt_num(v):
        if v is None or v == "": return ""
        return str(int(v)) if isinstance(v, float) and v == int(v) else str(v)

    def _size_str(key):
        hp = h_pitches.get(key)
        vp = v_pitches.get(key)
        if hp and vp:
            return f"{key[0]}x{key[1]}/H:{hp}/V:{vp}"
        elif hp:
            return f"{key[0]}x{key[1]}/{_fmt_num(hp)}"
        else:
            return f"{key[0]}x{key[1]}"

    # ── 晶片邊界 → 方向性 BUMP to IC edge ──────────────────────────────────
    chip_bbox = get_chip_bbox(top, scale) if (top is not None and scale is not None) else None

    def _edge_dists(cat_bumps_list):
        """回傳 (v_label, v_dist, h_dist)：垂直最近邊 + 水平最近側邊
        使用 real（實際最小距離），不用 avg（平均值）"""
        if not chip_bbox or not cat_bumps_list:
            return None, None, None
        xmin, ymin, xmax, ymax = chip_bbox
        # 找出到各邊的最小距離（使用 real，不用 avg）
        # 計算每個 bump 到各邊的距離
        d_top    = min(ymax - (b["cy"] + b["h"]/2) for b in cat_bumps_list)
        d_bottom = min((b["cy"] - b["h"]/2) - ymin  for b in cat_bumps_list)
        d_left   = min((b["cx"] - b["w"]/2) - xmin  for b in cat_bumps_list)
        d_right  = min(xmax - (b["cx"] + b["w"]/2) for b in cat_bumps_list)

        # 選擇最近的垂直邊
        if d_top <= d_bottom:
            v_label, v_dist = "BUMP to IC 上方edge", round(d_top, 3)
        else:
            v_label, v_dist = "BUMP to IC 下方edge", round(d_bottom, 3)
        # 選擇最近的水平邊
        h_dist = round(min(d_left, d_right), 3)

        # DEBUG: 驗證 BUMP to chip_edge 計算 (all distances use real min values, never avg)
        import sys as _sys_edge
        print(f"[EDGE_DISTS] count={len(cat_bumps_list)}, chip_bbox=({xmin:.1f}, {ymin:.1f}, {xmax:.1f}, {ymax:.1f})", file=_sys_edge.stderr)
        print(f"[EDGE_DISTS] all 4 edges: d_top={d_top:.2f}, d_bottom={d_bottom:.2f}, d_left={d_left:.2f}, d_right={d_right:.2f}", file=_sys_edge.stderr)
        print(f"[EDGE_DISTS] → selected: {v_label}={v_dist}um, side_edge={h_dist}um", file=_sys_edge.stderr)

        # 驗證：如果距離為負，表示 bump 延伸超出 chip 邊界
        if v_dist < 0 or h_dist < 0:
            print(f"[EDGE_DISTS] ⚠️  WARNING: negative distance detected (bump extends outside chip boundary)", file=_sys_edge.stderr)

        return v_label, v_dist, h_dist

    # ── Per-type background colours（與圖面 palette 一致）────────────────────
    _in_sorted  = sorted(input_types.keys())
    _out_sorted = sorted(output_types.keys())
    _dum_sorted = sorted(dummy_types.keys())
    _in_bg  = {k: _lighten(_PAL_INPUT [i%len(_PAL_INPUT )], 0.30) for i,k in enumerate(_in_sorted)}
    _out_bg = {k: _lighten(_PAL_OUTPUT[i%len(_PAL_OUTPUT)], 0.30) for i,k in enumerate(_out_sorted)}
    _dum_bg = {k: _lighten(_PAL_DUMMY [i%len(_PAL_DUMMY )], 0.30) for i,k in enumerate(_dum_sorted)}
    _in_edge_bg  = "FFFFFF"
    _out_edge_bg = "FFFFFF"
    _dum_edge_bg = "FFFFFF"

    # ── 預先計算 OUTPUT-to-INPUT 間距 ─────────────────────────────────────────
    _oi_gap = None
    if in_bumps_list and out_bumps_list:
        _in_top_y = max(b["cy"] + b["h"]/2 for b in in_bumps_list)
        _chip_center_x = (chip_bbox[0] + chip_bbox[2]) / 2 if chip_bbox else 0
        _chip_width = chip_bbox[2] - chip_bbox[0] if chip_bbox else 1000
        _near_out = [b for b in out_bumps_list
                     if abs(b["cx"] - _chip_center_x) <= _chip_width / 4]
        _out_bot_y = min(b["cy"] - b["h"]/2 for b in (_near_out or out_bumps_list))
        _oi_gap = round(abs(_in_top_y - _out_bot_y), 3)

    # ── V1.1 INPUT section ────────────────────────────────────────────────────
    # Group by H (all INPUT share same W); show each H as one row,
    # then edge/gap/W rows (A_{n+1}..A_{n+5})
    inp_by_h = {}
    for b in in_bumps_list:
        inp_by_h.setdefault(round(b["h"]), []).append(b)

    from collections import Counter as _Counter
    _inp_w_counts = _Counter(round(b["w"]) for b in in_bumps_list)
    inp_w_main = _inp_w_counts.most_common(1)[0][0] if _inp_w_counts else None
    _inp_w_all = sorted(_inp_w_counts)  # all unique W values (ascending)
    _inp_rep_key = (inp_w_main, min(inp_by_h)) if (inp_w_main and inp_by_h) else None
    _inp_h_pitch = h_pitches_in.get(_inp_rep_key)
    inp_h_gap = round(float(_inp_h_pitch) - float(inp_w_main), 3) if (_inp_h_pitch and inp_w_main) else None

    all_inp_range = _no_range_str([b["number"] for b in in_bumps_list])

    def _num_desc(base_num, desc):
        return f"{base_num} ({desc})" if base_num else f"({desc})"

    in_sym = 0
    for h_val in sorted(inp_by_h):
        # pick the bg color using the most common W for this H group
        _dominant_w = _Counter(round(b["w"]) for b in inp_by_h[h_val]).most_common(1)[0][0]
        bg = _in_bg.get((_dominant_w, h_val), "EBF3FB")
        nos = [b["number"] for b in inp_by_h[h_val]]
        in_sym += 1
        dat(ws, r + in_sym - 1, ISEC,   f"A{in_sym}",                             bg=bg, bold=True)
        dat(ws, r + in_sym - 1, ISEC+1, h_val,                                    bg=bg)
        dat(ws, r + in_sym - 1, ISEC+2, _num_desc(_no_range_str(nos), f"H={h_val}"), bg=bg)
        dat(ws, r + in_sym - 1, ISEC+3, "±2",                                     bg=bg)

    v_label_in, v_dist_in, h_dist_in = _edge_dists(in_bumps_list)
    _inp_hp_str = str(int(_inp_h_pitch)) if _inp_h_pitch and _inp_h_pitch == int(_inp_h_pitch) else str(_inp_h_pitch)
    for val, base_num, tol, desc in [
        (v_dist_in, all_inp_range, "±2", f"BUMP to IC {v_label_in.split('IC ')[-1]}" if v_label_in else "BUMP to IC edge"),
        (h_dist_in, "",           "±2", "BUMP to IC 側邊"),
        (inp_h_gap, "",           "±2", f"H-pitch gap = {_inp_hp_str}-{inp_w_main}" if (_inp_h_pitch and inp_w_main) else "H-pitch gap"),
        (_oi_gap,   "",           "",   "OUTPUT to INPUT gap"),
    ]:
        if val is None: continue
        in_sym += 1
        dat(ws, r + in_sym - 1, ISEC,   f"A{in_sym}",            bg=_in_edge_bg, bold=True)
        dat(ws, r + in_sym - 1, ISEC+1, val,                      bg=_in_edge_bg)
        dat(ws, r + in_sym - 1, ISEC+2, _num_desc(base_num, desc), bg=_in_edge_bg)
        dat(ws, r + in_sym - 1, ISEC+3, tol,                      bg=_in_edge_bg)
    # W 行：每個唯一 W 值各一行
    for w_val in _inp_w_all:
        in_sym += 1
        dat(ws, r + in_sym - 1, ISEC,   f"A{in_sym}",              bg=_in_edge_bg, bold=True)
        dat(ws, r + in_sym - 1, ISEC+1, w_val,                     bg=_in_edge_bg)
        dat(ws, r + in_sym - 1, ISEC+2, _num_desc("", "Bump W"),   bg=_in_edge_bg)
        dat(ws, r + in_sym - 1, ISEC+3, "±2",                      bg=_in_edge_bg)

    # ── V1.1 OUTPUT section ───────────────────────────────────────────────────
    # Each (W,H) type gets 2 rows: W row then H row; then edge/gap rows
    out_sym = 0
    for (out_w, out_h) in sorted(output_types):
        bg = _out_bg.get((out_w, out_h), "E2EFDA")
        nos = output_types[(out_w, out_h)]["nos"]
        out_sym += 1
        dat(ws, r + out_sym - 1, OSEC,   f"B{out_sym}",                                    bg=bg, bold=True)
        dat(ws, r + out_sym - 1, OSEC+1, out_w,                                            bg=bg)
        dat(ws, r + out_sym - 1, OSEC+2, _num_desc(_no_range_str(nos), "Bump W"),          bg=bg)
        dat(ws, r + out_sym - 1, OSEC+3, "±2",                                             bg=bg)
        out_sym += 1
        dat(ws, r + out_sym - 1, OSEC,   f"B{out_sym}",                                   bg=bg, bold=True)
        dat(ws, r + out_sym - 1, OSEC+1, out_h,                                           bg=bg)
        dat(ws, r + out_sym - 1, OSEC+2, _num_desc("", f"H={out_h}"),                     bg=bg)
        dat(ws, r + out_sym - 1, OSEC+3, "±2",                                            bg=bg)

    v_label_out, v_dist_out, h_dist_out = _edge_dists(out_bumps_list)
    _out_rep = sorted(output_types)[0] if output_types else None
    _out_hp = h_pitches_out.get(_out_rep) if _out_rep else None
    _out_vp = v_pitches_out.get(_out_rep) if _out_rep else None
    _out_h_gap = round(float(_out_hp) - float(_out_rep[0]), 3) if (_out_hp and _out_rep) else None
    _out_v_gap = round(float(_out_vp) - float(_out_rep[1]), 3) if (_out_vp and _out_rep) else None
    _out_hp_str = str(int(_out_hp)) if _out_hp and _out_hp == int(_out_hp) else str(_out_hp)
    _out_vp_str = str(int(_out_vp)) if _out_vp and _out_vp == int(_out_vp) else str(_out_vp)
    _v_edge_lbl = f"BUMP to IC {v_label_out.split('IC ')[-1]}" if v_label_out else "BUMP to IC edge"
    for val, tol, desc in [
        (v_dist_out, "±2", _v_edge_lbl),
        (h_dist_out, "±2", "BUMP to IC 側邊"),
        (_out_h_gap, "±2", f"H-pitch gap = {_out_hp_str}-{_out_rep[0]}" if (_out_hp and _out_rep) else "H-pitch gap"),
        (_out_v_gap, "±2", f"V-pitch gap = {_out_vp_str}-{_out_rep[1]}" if (_out_vp and _out_rep) else "V-pitch gap"),
    ]:
        if val is None: continue
        out_sym += 1
        dat(ws, r + out_sym - 1, OSEC,   f"B{out_sym}",             bg=_out_edge_bg, bold=True)
        dat(ws, r + out_sym - 1, OSEC+1, val,                        bg=_out_edge_bg)
        dat(ws, r + out_sym - 1, OSEC+2, _num_desc("", desc),        bg=_out_edge_bg)
        dat(ws, r + out_sym - 1, OSEC+3, tol,                        bg=_out_edge_bg)

    # ── V1.1 DUMMY section ────────────────────────────────────────────────────
    dum_sym = 0
    for (dum_w, dum_h) in sorted(dummy_types):
        bg = _dum_bg.get((dum_w, dum_h), "FFF2CC")
        nos = dummy_types[(dum_w, dum_h)]["nos"]
        dum_sym += 1
        dat(ws, r + dum_sym - 1, DSEC,   f"C{dum_sym}",                                        bg=bg, bold=True)
        dat(ws, r + dum_sym - 1, DSEC+1, dum_w,                                                bg=bg)
        dat(ws, r + dum_sym - 1, DSEC+2, _num_desc(_no_range_str(nos), f"{dum_w}×{dum_h} Bump W"), bg=bg)
        dat(ws, r + dum_sym - 1, DSEC+3, "±2",                                                 bg=bg)
        dum_sym += 1
        dat(ws, r + dum_sym - 1, DSEC,   f"C{dum_sym}",                                       bg=bg, bold=True)
        dat(ws, r + dum_sym - 1, DSEC+1, dum_h,                                               bg=bg)
        dat(ws, r + dum_sym - 1, DSEC+2, _num_desc("", f"{dum_w}×{dum_h} Bump H"),            bg=bg)
        dat(ws, r + dum_sym - 1, DSEC+3, "±2",                                                bg=bg)

    v_label_dum, v_dist_dum, h_dist_dum = _edge_dists(dum_bumps_list)
    _dv_lbl = f"BUMP to IC {v_label_dum.split('IC ')[-1]}" if v_label_dum else "BUMP to IC edge"
    for val, tol, desc in [
        (v_dist_dum, "±2", _dv_lbl),
        (h_dist_dum, "±2", "BUMP to IC 側邊"),
    ]:
        if val is None: continue
        dum_sym += 1
        dat(ws, r + dum_sym - 1, DSEC,   f"C{dum_sym}",        bg=_dum_edge_bg, bold=True)
        dat(ws, r + dum_sym - 1, DSEC+1, val,                   bg=_dum_edge_bg)
        dat(ws, r + dum_sym - 1, DSEC+2, _num_desc("", desc),   bg=_dum_edge_bg)
        dat(ws, r + dum_sym - 1, DSEC+3, tol,                   bg=_dum_edge_bg)

    # ── AMARK：gap-based cluster bbox（每個實體 mark 的外框尺寸）───────────
    amark_sizes = []
    if top is not None and scale is not None:
        _am_info = []
        for poly in top.polygons:
            if (poly.layer, poly.datatype) == AMARK_LAYER and len(poly.points) >= 3:
                pts = np.array(poly.points) * scale
                _am_info.append((float(pts[:,0].mean()), pts))
        _am_info.sort(key=lambda x: x[0])
        _am_clusts = []
        _AM_G = 500.0
        for _acx, _apts in _am_info:
            if _am_clusts and _acx - float(_am_clusts[-1][-1][0]) < _AM_G:
                _am_clusts[-1].append((_acx, _apts))
            else:
                _am_clusts.append([(_acx, _apts)])
        for _clust in _am_clusts:
            _all = np.vstack([p for _, p in _clust])
            aw = round(float(_all[:,0].max() - _all[:,0].min()), 2)
            ah = round(float(_all[:,1].max() - _all[:,1].min()), 2)
            if not (40 <= max(aw, ah) <= 2000):  # 排除雜散小 poly 和 chip outline 等超大 poly
                continue
            amark_sizes.append((aw, ah))
    if not amark_sizes:
        amark_sizes = [(100.0, 100.0)] * len(amarks)

    for i, (aw, ah) in enumerate(amark_sizes):
        dat(ws, r+i, ASEC,   f"D{i+1}",        bold=True)
        dat(ws, r+i, ASEC+1, f"{aw}x{ah}")
        dat(ws, r+i, ASEC+2, "±2")

    # ── AMARK 座標表（A1 A2 橫向並排）────────────────────────────────────────
    max_rows = max(in_sym, out_sym, dum_sym, len(amark_sizes), 1)
    r2 = r + max_rows + 3
    hdr(ws, r2, 1, "AMARK Position", bg="833C11")
    r2 += 1
    for i, (ax_pos, ay_pos) in enumerate(amarks, 1):
        base = (i - 1) * 3 + 1
        for j, h in enumerate(["Symbol", "X (um)", "Y (um)"], 0):
            hdr(ws, r2, base + j, h, bg="833C11")
    r2 += 1
    for i, (ax_pos, ay_pos) in enumerate(amarks, 1):
        base = (i - 1) * 3 + 1
        dat(ws, r2, base,     f"D{i}", bold=True)
        dat(ws, r2, base + 1, ax_pos, fmt="0.000")
        dat(ws, r2, base + 2, ay_pos, fmt="0.000")
    r2 += 1

    # ── Amark 放大圖（座標表下方）──
    img_start_row = r2 + 2
    for ri in range(img_start_row, img_start_row + 28):
        ws.row_dimensions[ri].height = 14
    if top is not None and scale is not None:
        for i, (ax_um, ay_um) in enumerate(amarks):
            png = render_amark_detail(top, scale, ax_um, ay_um, f"D{i+1}")
            _embed_image(ws, png, f"{get_column_letter(i * 3 + 1)}{img_start_row}",
                         width_px=440, height_px=440)

    # 欄寬 (ISEC=A=1, OSEC=E=5, DSEC=I=9, ASEC=M=13)
    # A=Symbol, B=WxH/Pitch, C=Number, D=Tol  (INPUT)
    # E=Symbol, F=WxH/Pitch/Vpitch, G=Number, H=Tol  (OUTPUT)
    # I=Symbol, J=WxH/Pitch, K=Number, L=Tol  (DUMMY)
    # M=Symbol, N=WxH, O=Tol  (AMARK)
    for col, w in [("A",28),("B",26),("C",40),("D",10),
                   ("E",28),("F",28),("G",40),("H",10),
                   ("I",28),("J",26),("K",40),("L",10),
                   ("M",13),("N",16),("O",10),
                   ("P",16),("Q",16),("R",16),("S",16)]:
        ws.column_dimensions[col].width = w

# ── Sheet 3：Bump Type Stat. ──────────────────────────────────────────────────
def write_bump_type_stat(ws, bumps):
    ws.title = "Bump Type Stat."
    # 統計每個 (W, H) 組合的數量，按數量降冪排列
    from collections import Counter
    counts = Counter((round(b["w"]), round(b["h"])) for b in bumps)
    rows = sorted(counts.items(), key=lambda x: -x[1])

    # Header
    for c, label in enumerate(["Bump Type", "Bump W", "Bump H", "Bump Amount"], 1):
        hdr(ws, 1, c, label, bg="1F4E79")

    for i, ((bw, bh), cnt) in enumerate(rows, 2):
        dat(ws, i, 1, i - 1)          # Type number
        dat(ws, i, 2, bw)
        dat(ws, i, 3, bh)
        dat(ws, i, 4, cnt)

    for col, w in [("A", 14), ("B", 12), ("C", 12), ("D", 14)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 20


# ── 主程式 ───────────────────────────────────────────────────────────────────
def main():
    print(f"Loading GDS: {GDS_PATH}")
    top, scale = load_flat(GDS_PATH)

    chip_w, chip_h = get_chip_size(top, scale)
    print(f"Chip: {chip_w} x {chip_h} um")

    bumps  = get_bumps(top, scale)
    amarks = get_amarks(top, scale)
    print(f"Bumps: {len(bumps)}, Amarks: {len(amarks)}")

    cats = [classify(b) for b in bumps]
    print(f"  INPUT:{cats.count('INPUT')}  OUTPUT:{cats.count('OUTPUT')}  DUMMY:{cats.count('DUMMY')}")
    print(f"  First 3: {bumps[:3]}")

    # Sheet 名稱：GDS 檔名（Excel 限制 31 字元）
    stem = Path(GDS_PATH).stem
    sheet_name = stem[:31]

    wb     = Workbook()
    ws_pad = wb.active
    ws_inf = wb.create_sheet("INFORMATION")

    write_pad_sheet(ws_pad, bumps, chip_w, chip_h, sheet_name)
    write_information(ws_inf, bumps, chip_w, chip_h, amarks, top=top, scale=scale)

    Path(OUT_PATH).parent.mkdir(exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")

def run(gds_path: str, out_path: str, mapping_json: str = None,
        original_stem: str = None, overrides: dict = None, window_um=None) -> str:
    """
    從外部呼叫的入口（供 gds_server.py 使用）。
    支援動態 layer mapping JSON 與空間分類（OUTPUT=密集行）。
    original_stem：原始 GDS 檔名（不含副檔名），用作 sheet1 名稱。
    window_um：簡圖各區域寬度（µm）。None = 依 bump pitch 自動計算，A2 Center 固定 ±1750。
    回傳輸出路徑。
    """
    import json as _json
    global BUMP_LAYERS, AMARK_LAYER, TEXT_NUM_LAYERS, TEXT_NAME_LAYER, CHIP_LAYER, classify

    # ── 自動偵測 mapping JSON 與 bump type overrides ──────────────────────────
    _gds_stem = Path(gds_path).stem
    _pname_auto = _gds_stem.split("_")[0] if "_" in _gds_stem else _gds_stem
    _out_parent = Path(out_path).parent if Path(out_path).suffix else Path(out_path)
    _auto_json_candidates = [
        _out_parent / f"{_pname_auto}.layer_mapping.json",
        Path(gds_path).parent / f"{_pname_auto}.layer_mapping.json",
        Path("output") / f"{_pname_auto}.layer_mapping.json",
    ]
    _auto_json_path = next((p for p in _auto_json_candidates if p.exists()), None)
    if _auto_json_path:
        import sys as _sys_auto
        if mapping_json is None:
            mapping_json = str(_auto_json_path)
            print(f"[AUTO JSON] 自動載入 {_auto_json_path.name}", file=_sys_auto.stderr)
        if overrides is None:
            try:
                _jdata = _json.loads(_auto_json_path.read_text(encoding="utf-8"))
                _bts = _jdata.get("bump_type_stats", [])
                if _bts:
                    overrides = {f"{int(b['w'])},{int(b['h'])}": b.get("cat", b.get("type", ""))
                                 for b in _bts if b.get("cat") or b.get("type")}
                    if overrides:
                        print(f"[AUTO OVERRIDES] {overrides}", file=_sys_auto.stderr)
            except Exception as _ex:
                print(f"[WARN] Auto override load: {_ex}", file=_sys_auto.stderr)

    # ── 從 mapping JSON 更新 layer 常數 ──────────────────────────────────────
    if mapping_json and Path(mapping_json).exists():
        _m = _json.loads(Path(mapping_json).read_text(encoding="utf-8")).get("mapping", {})
        def _first(key):
            lst = _m.get(key, [])
            return tuple(lst[0]) if lst else None

        # 注意：BUMP_LAYERS [(224,0)] 和 TEXT_NUM_LAYERS [(224,0), (225,0)] 不被 mapping JSON 覆蓋
        # Layer 224/225 DType 0：根據內容強制分類
        #   - 含 #POLY（polygon）一律為 BUMP/PAD
        #   - 含 #TEXT 一律為 Bump Text (序號)
        if _first("amark"):           AMARK_LAYER     = _first("amark")
        if _first("bump_text_name"):  TEXT_NAME_LAYER = _first("bump_text_name")
        if _first("chip_size"):       CHIP_LAYER      = _first("chip_size")

    # ── 載入 GDS ─────────────────────────────────────────────────────────────
    top, scale = load_flat(gds_path)
    chip_w, chip_h = get_chip_size(top, scale)
    bumps  = get_bumps(top, scale)
    amarks = get_amarks(top, scale)

    # 檢查坐標尺度是否合理：如果 bump 尺寸在 10-100 但坐標在 1-100，
    # 表示坐標小了 10 倍，需要重新縮放所有坐標
    if bumps:
        # 使用 real（實際最大值），不用 avg（平均值）
        max_bump_size = max(max(b["w"], b["h"]) for b in bumps)
        max_coord = max(max(abs(b["cx"]), abs(b["cy"])) for b in bumps)

        import sys as _sys_scale_check
        print(f"\n[SCALE CHECK] max_bump_size={max_bump_size:.2f}, max_coord={max_coord:.2f}, scale={scale}", file=_sys_scale_check.stderr)

        # 如果 bump 尺寸 > 10 但 max_coord < 100，坐標小了 10 倍
        if max_bump_size > 10 and max_coord < 100:
            print(f"[SCALE ADJUST] Detected 10x scaling, adjusting all coordinates", file=_sys_scale_check.stderr)
            # 重新縮放所有坐標、芯片尺寸、alignment marks
            for b in bumps:
                b["cx"] *= 10
                b["cy"] *= 10
                b["w"] *= 10
                b["h"] *= 10
            if chip_w and chip_h:
                chip_w *= 10
                chip_h *= 10
            for a in amarks:
                a["cx"] *= 10
                a["cy"] *= 10
                a["w"] *= 10
                a["h"] *= 10

    # ── 空間分類（密集單排 + 晶片下半 → INPUT，上半 → OUTPUT，其餘 → DUMMY）──
    _chip_bbox = get_chip_bbox(top, scale)
    _chip_cy   = (_chip_bbox[1] + _chip_bbox[3]) / 2 if _chip_bbox else 0

    import sys as _sys_spatial
    print(f"\n[SPATIAL CLASS] chip_bbox={_chip_bbox}, chip_cy={_chip_cy:.2f}", file=_sys_spatial.stderr)

    _y_grps: dict = {}
    for _b in bumps:
        _ky = round(_b["cy"] * 2) / 2   # ±0.5 μm 容差，同 Y 行歸為同組
        _y_grps.setdefault(_ky, []).append(id(_b))

    # INPUT：密集單排 + 晶片下半部（固定 1 排）
    _input_ids  = {i for ky, ids in _y_grps.items() if len(ids) > 10 and ky < _chip_cy for i in ids}
    # INPUT 上緣 Y（所有 INPUT bump 的最高點）
    _input_bumps = [b for b in bumps if id(b) in _input_ids]
    _input_top_y = max(b["cy"] + b["h"] / 2 for b in _input_bumps) if _input_bumps else _chip_cy
    # OUTPUT：INPUT 上緣以上的密集行（1 排以上）
    _output_ids = {i for ky, ids in _y_grps.items() if len(ids) > 10 and ky > _input_top_y for i in ids}

    print(f"[SPATIAL CLASS] input_count={len(_input_ids)}, input_top_y={_input_top_y:.2f}", file=_sys_spatial.stderr)
    print(f"[SPATIAL CLASS] output_count={len(_output_ids)}", file=_sys_spatial.stderr)

    # 分類結果直接寫入 bump dict，避免 id() 跨函式失效
    for _b in bumps:
        if id(_b) in _input_ids:
            _b["_cat"] = "INPUT"
        elif id(_b) in _output_ids:
            _b["_cat"] = "OUTPUT"
        else:
            _b["_cat"] = "DUMMY"

    # 套用使用者覆寫（GUI Bump Type 分類確認表）
    if overrides:
        import sys as _sys
        print(f"\n{'='*60}", file=_sys.stderr)
        print(f"[OVERRIDE] 開始套用 Bump Type 覆寫", file=_sys.stderr)
        print(f"[OVERRIDE] 收到 {len(overrides)} 個覆寫規則", file=_sys.stderr)

        _matched = 0
        _matched_by_type = {"INPUT": 0, "OUTPUT": 0, "DUMMY": 0}

        # 建立容差匹配表（处理浮点精度差异）
        _ovr_lookup = {}
        for _key_str, _cat in overrides.items():
            try:
                _w_str, _h_str = _key_str.split(',')
                _w_float = float(_w_str)
                _h_float = float(_h_str)
                _ovr_lookup[(round(_w_float, 4), round(_h_float, 4))] = _cat
                print(f"[OVERRIDE] 規則: ({_w_float}, {_h_float}) → {_cat}", file=_sys.stderr)
            except:
                pass

        print(f"[OVERRIDE] 準備匹配 {len(bumps)} 個 bumps...", file=_sys.stderr)

        for _b in bumps:
            _w_rounded = round(_b['w'], 4)
            _h_rounded = round(_b['h'], 4)
            _key = (_w_rounded, _h_rounded)
            if _key in _ovr_lookup:
                _old_cat = _b.get("_cat", "DUMMY")
                _new_cat = _ovr_lookup[_key]
                _b["_cat"] = _new_cat
                _matched += 1
                _matched_by_type[_new_cat] += 1

                # 首次匹配时输出示例
                if _matched <= 5:
                    print(f"[OVERRIDE] 匹配: ({_w_rounded},{_h_rounded}) {_old_cat} → {_new_cat}", file=_sys.stderr)

        print(f"[OVERRIDE] 完成匹配: {_matched}/{len(bumps)} bumps ({_matched_by_type})", file=_sys.stderr)
        print(f"{'='*60}\n", file=_sys.stderr)

    # 覆寫全局 classify 以反映 overrides 結果
    global classify
    def _classify_bumps(b):
        return b.get("_cat", "DUMMY")
    # 替代全局 classify，使 render_key_sections 能看到 overrides
    classify = _classify_bumps

    # ── 自動計算 window_um（若未指定）────────────────────────────────────────
    if window_um is None:
        window_um = compute_auto_window_um(bumps, chip_w or 30000)
        print(f"[AUTO WINDOW] window_um={window_um}")

    # ── 產生 Excel ───────────────────────────────────────────────────────────
    stem = original_stem if original_stem else Path(gds_path).stem
    wb     = Workbook()
    ws_pad = wb.active
    ws_inf = wb.create_sheet("INFORMATION")
    write_pad_sheet(ws_pad, bumps, chip_w, chip_h, stem[:31])
    write_information(ws_inf, bumps, chip_w, chip_h, amarks, top=top, scale=scale, window_um=window_um)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path


if __name__ == "__main__":
    main()
